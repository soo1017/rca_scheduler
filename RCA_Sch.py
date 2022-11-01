####################### Import
import tkinter as tk
from tkinter.constants import END, FALSE, LEFT, TRUE
from tkinter.filedialog import askopenfilename
import random

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

import re
import operator
import numpy as np

import os
from twilio.rest import Client
################################


# Class of Widget - Label
class label_gui():
    def __init__(self, master):
        self.label = tk.Label(master)

    def setLabel(self, textS, fontS, padyS):
        self.label.config(text=textS)
        self.label.config(font=fontS)
        self.label.config(pady=padyS)

    def setLabel0(self, textS, fontS, padxS, padyS, justifyS, borderwidthS, fgS):
        self.label.config(text=textS)
        self.label.config(font=fontS)
        self.label.config(padx=padxS)
        self.label.config(pady=padyS)
        self.label.config(justify=justifyS)
        self.label.config(borderwidth=borderwidthS)
        self.label.config(fg=fgS)
    
    def setLabelImage(self, imageS):
        self.label.config(image=imageS)

    def setFontColor(self, fgS):
        self.label.config(fg=fgS)

    def setGridLabel(self, columnS, rowS, columnspanS):
        self.label.grid(column=columnS, row=rowS, columnspan=columnspanS)
    
    def setGridLabel0(self, stickyS, columnS, rowS, columnspanS):
        self.label.grid(sticky=stickyS, column=columnS, row=rowS, columnspan=columnspanS)

    def setGridLabelImage(self, columnS, rowS, rowspanS):
        self.label.grid(column=columnS, row=rowS, rowspan=rowspanS)

    def configurePadx(self, padxS):
        self.label.config(padx=padxS)

    def setSticky(self, stickyS):
        self.label.grid(sticky=stickyS)

    def setPadx(self, padxS):
        self.label.grid(padx=padxS)

    def configureText(self, textS):
        self.label.config(text=textS)

    def setJustify(self, justifyS):
        self.label.config(justify=justifyS)

# Class of Widget - Canvas
class canvas_gui():
    def __init__(self, master):
        self.canvas = tk.Canvas(master)

    def createLine(self, x0, y0, x1, y1):
        self.canvas.create_line(x0, y0, x1, y1)

    def setLine(self, widthS, heightS):
        self.canvas.config(width=widthS)
        self.canvas.config(height=heightS)

    def setGridLine(self, columnS, rowS, columnspanS):
        self.canvas.grid(column=columnS, row=rowS, columnspan=columnspanS)

# Class of Widget - Entry
class entry_gui():
    def __init__(self, master):
        self.entry = tk.Entry(master)
    
    def setEntry(self, textS, widthS, borderwidthS, bgS):
        self.entry.config(text=textS)
        self.entry.config(width=widthS)
        self.entry.config(borderwidth=borderwidthS)
        self.entry.config(bg=bgS)

    def setGridEntry(self, columnS, rowS, columnspanS, padyS):
        self.entry.grid(column=columnS, row=rowS, columnspan=columnspanS, pady=padyS)

    def deleteValueEntry(self):
        self.entry.delete(0, END)

    def setValueEntry(self, textS):
        self.entry.insert(0, textS)

    def setSticky(self, stickyS):
        self.entry.grid(sticky=stickyS)

    def setPadx(self, padxS):
        self.entry.grid(padx=padxS)

    def setJustify(self, justifyS):
        self.entry.config(justify=justifyS)

    def setShow(self, showS):
        self.entry.config(show=showS)

    def bindClick(self):
        self.entry.bind("<Button-1>", self.clearTextEntry)

    def clearTextEntry(self, event): # note that you must include the event as an arg, even if you don't use it.
        self.entry.delete(0, "end")
        return None

    def getText(self):
        return self.entry.get()

# Class of Widget - Button
class button_gui():
    def __init__(self, master):
        self.button = tk.Button(master)
    
    def setButton(self, textS, widthS, bgS, bdS, fgS, commandS):
        self.button.config(text=textS)
        self.button.config(width=widthS)
        self.button.config(bg=bgS)
        self.button.config(bd=bdS)
        self.button.config(fg=fgS)
        self.button.config(command=commandS)

    def setGridButton(self, columnS, rowS, columnspanS, padyS):
        self.button.grid(column=columnS, row=rowS, columnspan=columnspanS, pady=padyS)

    def setSticky(self, stickyS):
        self.button.grid(sticky=stickyS)

    def setPadx(self, padxS):
        self.button.grid(padx=padxS)
    
    def setPady(self, padyS):
        self.button.grid(pady=padyS)

    def getState(self):
        return self.button['state']

    def setState(self, stateS):
        self.button.configure(state=stateS)

# Class of Widget - Radio Button
class radiobutton_gui():
    def __init__(self, master):
        self.radiobutton = tk.Radiobutton(master)
    
    def setRadioButton(self, textS, variableS, valueS):
        self.radiobutton.config(text=textS)
        self.radiobutton.config(variable=variableS)
        self.radiobutton.config(value=valueS)

    def setGridRadioButton(self, columnS, rowS, columnspanS):
        self.radiobutton.grid(column=columnS, row=rowS, columnspan=columnspanS)

    def setStickyRadioButton(self, stickyS):
        self.radiobutton.grid(sticky=stickyS)

    def setPadxRadioButton(self, padxS):
        self.radiobutton.grid(padx=padxS)
    
    def setPadyRadioButton(self, padyS):
        self.radiobutton.grid(pady=padyS)

# Class of Widget - Message
class message_gui():
    def __init__(self, master):
        self.message = tk.Message(master)

    def setMessage(self, textS, fontS, fgS, widthS, padyS):
        self.message.config(text=textS)
        self.message.config(font=fontS)
        self.message.config(fg=fgS)
        self.message.config(width=widthS)
        self.message.config(pady=padyS)

    def setGridMessage(self, columnS, rowS, columnspanS):
        self.message.grid(column=columnS, row=rowS, columnspan=columnspanS)

    def changeText(self, textS):
        self.message.config(text=textS)

    def changeColor(self, colorS):
        self.message.config(fg=colorS)



#
################################### Class of 'Excel' ############################
#

############################################# Class 'Excel'
class Excel():
    RCA_Name=[]
    RCA_Phone=[]
    RCA_Value=[]
    RCA_PreviousRoute_Info=[]
    Supervisor={}
    Route_Name=[]
    Route_Value=[]
    Route_Value_LeftOver = []
    New_Sch=[]
    Sch_1stRead=[]
    Sch_2ndRead_AfterModifying=[]
    Route_Insertion_Rate = 0
    Cell_Col_Width = 14
    Mode_One_OffDay_Req = 'No'


    def __init__(self):
        self = self

    ############################################################ for SetUp
    # Class Function to set 'load_workbook'
    def set_Wb(self, wbS):
        self.wb = wbS

    ############################################################ for SetUp
    # Class Function to get 'load_workbook'
    def get_Wb(self):
        return self.wb

    ############################################################ for SetUp
    # Class Function to set Supervisor Phone # from setup window
    def Set_Supervisor_Phone(self, phoneE):
        if phoneE is not None: 
            self.Supervisor['Phone'] = phoneE

    # Class Function to get Supervisor Phone # from setup window
    def Get_Supervisor_Phone(self):
        return self.Supervisor.get('Phone')
    
    #################################################################### for SetUp
    # Class Function to set a new Route Insertion Rate from setup window
    def Set_New_Route_Insertion_Rate(self, newRTEInsertRateE):
        temp_Int0_RTEInsertRate = int(newRTEInsertRateE)
        if newRTEInsertRateE != 0 and (temp_Int0_RTEInsertRate > 0 and temp_Int0_RTEInsertRate <= 100): 
            self.Route_Insertion_Rate = temp_Int0_RTEInsertRate
        else:
            self.Route_Insertion_Rate = 0

    # Class Function to get a new Route Insertion Rate from setup window
    def Get_New_Route_Insertion_Rate(self):
        return self.Route_Insertion_Rate

    #################################################################### for SetUp
    # Class Function to set a new Cell Column Width from setup window
    def Set_New_Cell_Column_Width(self, newCellColWidth):
        temp_Int0_CellColWidth = int(newCellColWidth)
        if int(newCellColWidth) != 0 and temp_Int0_CellColWidth > 6:
            self.Cell_Col_Width = int(newCellColWidth)
        else:
            self.Cell_Col_Width = 16
    
    # Class Function to get a new Cell Column Width from setup window
    def Get_New_Cell_Column_Width(self):
        return self.Cell_Col_Width

    #################################################################### for SetUp
    # Class Function to set a varible of 'Mode_One_OffDay_Req'
    def Set_One_OffDay_Required(self, mode_One_OffDay_Required):
        self.Mode_One_OffDay_Req = mode_One_OffDay_Required

    # Class Function to get a varible of 'Mode_One_OffDay_Req'
    def Get_One_OffDay_Required(self):
        return self.Mode_One_OffDay_Req

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'RCA_Value'
    def Empty_A_Variable_RCA_Value(self):
        self.RCA_Value.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'RCA_Name'
    def Empty_A_Variable_RCA_Name(self):
        self.RCA_Name.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'RCA_Phone'
    def Empty_A_Variable_RCA_Phone(self):
        self.RCA_Phone.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'RCA_PreviousRoute_Info'
    def Empty_A_Variable_RCA_PreviousRoute_Info(self):
        self.RCA_PreviousRoute_Info.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'Route_Name'
    def Empty_A_Variable_Route_Name(self):
        self.Route_Name.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'Route_Value'
    def Empty_A_Variable_Route_Value(self):
        self.Route_Value.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'Sch_1stRead'
    def Empty_A_Variable_Sch_1stRead(self):
        self.Sch_1stRead.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'Sch_2ndRead_AfterModifying'
    def Empty_A_Variable_Sch_2ndRead_AfterModifying(self):
        self.Sch_2ndRead_AfterModifying.clear()

    ################################################ for Variable SetUp
    # Class Function to empty a variable 'New_Sch'
    def Empty_A_Variable_New_Sch(self):
        self.New_Sch.clear()

    ################################################ for Variable SetUp
    # Class Function to empty all variables
    def Empty_All_Variables(self):
        self.RCA_Name.clear()
        self.RCA_Phone.clear()
        self.RCA_Value.clear()
        self.RCA_PreviousRoute_Info.clear()
        self.Supervisor.clear()
        self.Route_Name.clear()
        self.Route_Value.clear()
        self.Route_Value_LeftOver.clear()
        self.New_Sch.clear()
        self.Sch_1stRead.clear()
        self.Sch_2ndRead_AfterModifying.clear()

    ################################################################## for Supports
    ###############################################################################
    # 1. a (class) function to sort Dic by is values (with descending order)
    def sort_Dic_ByValue(self, dict01):
        return dict( sorted(dict01.items(), key=operator.itemgetter(1), reverse=True))

    # 2. a (class) function to verify if a sheet is created or not
    def Verify_A_Sheet_Existed_Or_After_Created(self, sheet_Name0, excel_File0):
        if excel_File0 is not None:                                 # Check After a sheet is created
            temp_WB = load_workbook(excel_File0)
        else:                                                        # Check a existing sheet
            temp_WB = self.get_Wb()
        sheet_Name_WB = str(sheet_Name0).strip().replace('/', '.').lower()
        flag_Sheet_Name = FALSE
        if sheet_Name_WB in temp_WB.sheetnames:
            flag_Sheet_Name = TRUE
        return flag_Sheet_Name

    # 3. Class Function to check if there is any 'L' or 'Lv' left in Route_Value
    def Check_L_Lv_Route_Value(self):
        temp_list_Day = ['Sat', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        flag_Still_L = FALSE
        for dd in range(0, len(self.Route_Value)):
            for keyDD, valueDD in self.Route_Value[dd]:
                if keyDD in temp_list_Day:
                    if valueDD == 'L' or valueDD == 'Lv':
                        flag_Still_L = TRUE
                        break
            if flag_Still_L:
                break

        return flag_Still_L

    # 4. Class Function to convert day into number
    def Check_Day_2Num(self, offDayC):
        if offDayC == 'sat':
            return 0
        elif offDayC == 'mon':
            return 1
        elif offDayC == 'tue':
            return 2
        elif offDayC == 'wed':
            return 3
        elif offDayC == 'thu':
            return 4
        else:
            return 5

    # 5. Class Function to convert num into day
    def Check_Num_2Day(self, numS):
        if numS == 0:
            return 'Sat'
        elif numS == 1:
            return 'Mon'
        elif numS == 2:
            return 'Tue'
        elif numS == 3:
            return 'Wed'
        elif numS == 4:
            return 'Thu'
        elif numS == 5:
            return 'Fri'
    
    # 6. Class Function to pick any random number among a ceratin numbers
    def Pick_A_Random_Number(self, list_OffDays_Regular):                                     # An argument of array of "Off-days" like ['Wed', 'Thu']
        list_Num = []
        for i in range(0, len(list_OffDays_Regular)):
            list_Num.append(self.Check_Day_2Num(list_OffDays_Regular[i].lower()))
        flag_Match = TRUE
        A_random_num = 0
        while flag_Match:
            A_random_num = random.randint(0, 6)
            if A_random_num in list_Num:
                flag_Match = TRUE
            else:
                flag_Match = FALSE

        return A_random_num

    # 7. Class Function to Convert a Sheet Name into List of 1 Week's Days
    def Convert_Sheet_Name_2Week_Days(self, Sheet_Name0):

        new_Sheet_Name0 = str(Sheet_Name0).strip().split('(')
        pp_Wk0 = new_Sheet_Name0[0].replace('w', '')
        pp_Wk1 = pp_Wk0.split('.')
        legend1 = 'PP ' + pp_Wk1[0] + ' Wk ' + pp_Wk1[1] + '\n' + 'RCA\'s Name'

        cell_Days_AWeek = [legend1]
        list_Days_AWeek = ['Sat', 'Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri']

        temp_List_Days0 = Sheet_Name0.split('(')
        temp_List_Days1 = temp_List_Days0[1].split(')')
        temp_List_Days2 = temp_List_Days1[0].split('-')

        if '/' in temp_List_Days2[0] and '/' in temp_List_Days2[1]:
            temp_Start_Date = temp_List_Days2[0].split('/')             # [MM, DD, YY]
            temp_End_Date = temp_List_Days2[1].split('/')               # [MM, DD, YY]
        else:
            temp_Start_Date = temp_List_Days2[0].split('.')             # [MM, DD, YY]
            temp_End_Date = temp_List_Days2[1].split('.')               # [MM, DD, YY]

        # Convert string into integer
        temp_Start_Date = list(map(int, temp_Start_Date))
        temp_End_Date = list(map(int, temp_End_Date))
        flag_Month_Change = FALSE
        flag_Year_Change = FALSE
        if temp_Start_Date[1] != temp_End_Date[1]:
            flag_Month_Change = TRUE
        if temp_Start_Date[2] != temp_End_Date[2]:
            flag_Year_Change = TRUE

        run_Date = temp_Start_Date[1]
        run_Month = temp_Start_Date[0]
        run_Year = temp_Start_Date[2]

        for i in range(7):
            if i != 0:
                run_Date = run_Date + 1

            if flag_Month_Change:                               # Month Changed
                if flag_Year_Change:                            # Month & Year Changed
                    if run_Date == 31:
                        cell_Days_AWeek.append(list_Days_AWeek[i] + "\n (" + str(run_Month) + '/' + str(run_Date) + '/' + str(run_Year) + ')')
                        run_Month = 1
                        run_Date = 0
                        run_Year = run_Year + 1
                    else:
                        cell_Days_AWeek.append(list_Days_AWeek[i] + "\n (" + str(run_Month) + '/' + str(run_Date) + '/' +str(run_Year) + ')')
                else:                                           # Only Month Changed
                    if ((run_Date == 29 and temp_Start_Date[0] == 2 and self.isAMultipleOf4(temp_Start_Date[2]))
                        or(run_Date == 28 and temp_Start_Date[0] == 2 and not self.isAMultipleOf4(temp_Start_Date[2]))
                            or(run_Date == 30 and (temp_Start_Date[0] == 4 or temp_Start_Date[0] == 6 or temp_Start_Date[0] == 9 or temp_Start_Date[0] == 11))
                                or(run_Date == 31)):          
                        cell_Days_AWeek.append(list_Days_AWeek[i] + "\n (" + str(run_Month) + '/' + str(run_Date) + '/' + str(run_Year) + ')')
                        run_Month = run_Month + 1
                        run_Date = 0
                    else:
                        cell_Days_AWeek.append(list_Days_AWeek[i] + "\n (" + str(run_Month) + '/' + str(run_Date) + '/' + str(run_Year) + ')')
            else:                                               # Month & Year Not Changed
                cell_Days_AWeek.append(list_Days_AWeek[i] + "\n (" + str(run_Month) + '/' + str(run_Date) + '/' + str(run_Year) + ')')

        cell_Days_AWeek.pop(2)
        return cell_Days_AWeek
    
    # 8. Check an integer if it is a multiple of 4
    def isAMultipleOf4(self, n):
        if int(n) % 4 == 0:
            return TRUE
        else:
            return FALSE


    ###############################################################################


    ################################################################################ for GEN & MOD
    # Class Function to extract a necessary table content by reading a sheet content
    def Extract_Table_Content_By_Reading_A_Sheet(self, sheet_NameE, starting_LegendE, excel_Reading):
        flag_RCA_Schedule = FALSE
        temp_Sheet_Name = str(sheet_NameE).strip().lower()
        # Handling a Sheet of RCA Schedule
        if 'ww' in temp_Sheet_Name:
            flag_RCA_Schedule = TRUE
        temp_Starting_LegendE = str(starting_LegendE).strip().lower()
        # temp_WB
        new_list = []
        if excel_Reading is not None:
            temp_WB = load_workbook(excel_Reading)
            new_list = [str(item).lower() for item in temp_WB.sheetnames]
        else:
            new_list = [str(item).lower() for item in self.get_Wb().sheetnames]
    
        try :
            new_list.index(temp_Sheet_Name)
            if excel_Reading is not None:
                sheet_Interested = temp_WB[temp_WB.sheetnames[new_list.index(temp_Sheet_Name)]]
            else:
                sheet_Interested = self.get_Wb()[self.get_Wb().sheetnames[new_list.index(temp_Sheet_Name)]]
            table_Interested = sheet_Interested.iter_rows()
            filtered_Table_Interested = list(([cell.value if cell.value is not None else "" for cell in row] for row in table_Interested))

            # Find the size of really necessary portion of the RCA table
            start_Col = 0
            start_Row = 0
            end_Col = 0
            end_Row = 0
            flag_End = FALSE
        
            for i in range(len(filtered_Table_Interested)):             # Row Interation
                for j in range(len(filtered_Table_Interested[i])):      # Col Interation
                    if (str(filtered_Table_Interested[i][j]).strip().lower() == temp_Starting_LegendE 
                        or temp_Starting_LegendE in str(filtered_Table_Interested[i][j]).strip().lower()):
                        start_Row = i
                        start_Col = j

                        flag_ROW = FALSE
                        Counter_Space = 0
                        for ii in range(i+1, len(filtered_Table_Interested)):       ## Row-wise Checking
                            if not flag_RCA_Schedule:                           ## with 'RCA' or 'Route' Sheet
                                if filtered_Table_Interested[ii][j] == '':
                                    end_Row = ii
                                    flag_ROW = TRUE
                                    break
                            else:                                               ## with a RCA Schedule Sheet
                                if filtered_Table_Interested[ii][j] == '':
                                    if Counter_Space == 0:
                                        Counter_Space += 1
                                    else:
                                        end_Row = ii
                                        flag_ROW = TRUE
                                        break
                        if not flag_ROW:
                            end_Row = len(filtered_Table_Interested)

                        flag_COL = FALSE
                        for jj in range(j+1, len(filtered_Table_Interested[0])):    ## Column-wise Checking
                            if not flag_RCA_Schedule:                           ## with 'RCA' or 'Route' Sheet
                                if filtered_Table_Interested[i][jj] == '':
                                    end_Col = jj
                                    flag_COL = TRUE
                                    break
                            else:                                               ## with a RCA Schedule Sheet
                                if str(filtered_Table_Interested[i][jj]).strip().lower() == 'comments':
                                    end_Col = jj
                                    flag_COL = TRUE
                                    break
                        if not flag_COL:
                            end_Col = len(filtered_Table_Interested[0])
                        flag_End = TRUE
                        break
                if flag_End:
                    break
            # Extract only part of necessary info from the original 2D list
            temp_Array = np.array(filtered_Table_Interested)
            fine_Table_Interested_Array = temp_Array[start_Row:end_Row, start_Col:end_Col]
            fine_Table_Interested_List = fine_Table_Interested_Array.tolist()

            return [TRUE, fine_Table_Interested_List]
        except ValueError:
            return [FALSE, None]

    ##################################################### for GEN
    ##### Calling 'Extract_Table_Content_By_Reading_A_Sheet'
    # Class Function to read in the contents of RCA table
    def Read_RCA(self):
        # Calling the function 'Extract_Table_Content_By_Reading_A_Sheet'
        if not self.Extract_Table_Content_By_Reading_A_Sheet('rca', 'rca\'s name', None)[0]:          ## Failed - FALSE
            RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t read 'RCA' sheet")
            RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
            return FALSE
        else:                                                                                   ## TRUE
            fine_RCA_List = self.Extract_Table_Content_By_Reading_A_Sheet('rca', 'rca\'s name', None)[1]

            new_list1 = [str(item1).lower() for item1 in fine_RCA_List[0]]
            index_4Phone = new_list1.index('phone')                             
            index_4RCA_Name = new_list1.index('rca\'s name')                        
            if index_4Phone == -1 or index_4RCA_Name == -1:
                RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t find a legend 'Phone' or 'RCA\'s Name")
                RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
                return FALSE

            # Make RCA_Value [{'Phone': '123-456-7890', ... 'Report Hour': '9am'}, {}], RCA_Name(above), and RCA_Phone based on what is found above
            for n in range(1, len(fine_RCA_List)):                  # Row-wise Iteration
                temp_Dic_RCA = {}
                for m in range(0, len(fine_RCA_List[0])):           # Column-wise Iteration
                    if m >= 1:                                      # from 'Phone' legend (index = 1)
                        if fine_RCA_List[n][m] != '':
                            temp_Dic_RCA[fine_RCA_List[0][m]] = fine_RCA_List[n][m]
                        else:
                            temp_Dic_RCA[fine_RCA_List[0][m]] = None
                    if m == index_4Phone:
                        if fine_RCA_List[n][m] == '':
                            self.RCA_Phone.append(None)
                        else:
                            self.RCA_Phone.append(str(fine_RCA_List[n][m]).strip())
                    if m == index_4RCA_Name:
                        self.RCA_Name.append(str(fine_RCA_List[n][m]))

                self.RCA_Value.append(temp_Dic_RCA)

            return TRUE

    ####################################################### for GEN
    ##### Calling 'Extract_Table_Content_By_Reading_A_Sheet'
    # Class Function to read in the contents of Route table
    def Read_Route(self):
        fine_RCA_Route_List = []
        # Calling the function 'Extract_Table_Content_By_Reading_A_Sheet'
        if not self.Extract_Table_Content_By_Reading_A_Sheet('route', 'route', None)[0]:          ## Failed - FALSE
            RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t read a 'Route' sheet")
            RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
            return FALSE
        else:                                                                                   ## TRUE
            fine_RCA_Route_List = self.Extract_Table_Content_By_Reading_A_Sheet('route', 'route', None)[1]

            new_list1 = [str(item1).lower() for item1 in fine_RCA_Route_List[0]]
            index_4Route = new_list1.index('route')                    
            if index_4Route == -1:
                RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t find a legend 'Route'")
                RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
                return FALSE

            # Make Route_Value [{'REG's Name': 'Sarang', ... 'Fri': None}, {}], Route_Name(above) based on what is found above
            for n in range(1, len(fine_RCA_Route_List)):
                temp_Dic_Route = {}
                for m in range(0, len(fine_RCA_Route_List[0])):
                    if m >= 1:
                        if fine_RCA_Route_List[n][m] != '':
                            if "REG\'s Name" in fine_RCA_Route_List[0][m] or "Station" in fine_RCA_Route_List[0][m]:
                                temp_Dic_Route[fine_RCA_Route_List[0][m]] = str(fine_RCA_Route_List[n][m]).strip()
                            else:
                                temp_Dic_Route[fine_RCA_Route_List[0][m]] = str(fine_RCA_Route_List[n][m]).strip().capitalize()
                        else:
                            temp_Dic_Route[fine_RCA_Route_List[0][m]] = None
                    if m == index_4Route:
                        self.Route_Name.append(fine_RCA_Route_List[n][m])
                            
                self.Route_Value.append(temp_Dic_Route)

            return TRUE
        
    ######################################################################### for GEN & MOD
    ##### Calling 'Extract_Table_Content_By_Reading_A_Sheet'
    # Class Function to read in the content of RCA's previously made schedule
    def Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(self, sheet_Name01, excel_Name01):
        a_RCA_Schedule_List = []
        new_Sheet_Name01 = str(sheet_Name01).strip().replace('/', '.').lower()
        fine_SchValue_AllRCA_List = []
        # Calling the function 'Extract_Table_Content_By_Reading_A_Sheet'
        if (not self.Extract_Table_Content_By_Reading_A_Sheet(new_Sheet_Name01, 'rca\'s name', None)[0]
            or not self.Extract_Table_Content_By_Reading_A_Sheet(new_Sheet_Name01, 'rca\'s name', excel_Name01)[0]):          ## Failed - FALSE
            RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t read a ", new_Sheet_Name01, ' sheet') 
            RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
            return [FALSE, None]
        else:
            if excel_Name01 is not None:                                                                                               ## TRUE
                a_RCA_Schedule_List = self.Extract_Table_Content_By_Reading_A_Sheet(new_Sheet_Name01, 'rca\'s name', excel_Name01)[1]
            else:
                a_RCA_Schedule_List = self.Extract_Table_Content_By_Reading_A_Sheet(new_Sheet_Name01, 'rca\'s name', None)[1]
            for i1 in range(2, len(a_RCA_Schedule_List)):
                temp_SchValue_OneRCA = []
                for j1 in range(0, len(a_RCA_Schedule_List[i1])):
                    if a_RCA_Schedule_List[i1][j1] != '':
                        temp_SchValue_OneRCA.append(a_RCA_Schedule_List[i1][j1])

                fine_SchValue_AllRCA_List.append(temp_SchValue_OneRCA)
    
            return [TRUE, fine_SchValue_AllRCA_List]

    ############################################################################################### for MOD
    ##### Calling 'Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule'
    # Class Function to read in the content of RCA's previously made schedule as 1st for comparison
    def Read_A_Sheet_1st(self, sheet_Name1):
        temp_Sheet_Name1 = str(sheet_Name1).strip().lower()
        if self.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(temp_Sheet_Name1, None)[0]:                          
            temp_2Array1 = self.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(temp_Sheet_Name1, None)[1]
        else:
            RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t read a ", temp_Sheet_Name1, ' sheet') 
            RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
        for inner_list1 in temp_2Array1:
            self.Sch_1stRead.append(inner_list1)

        print("self.Sch_1stRead: ", self.Sch_1stRead)

    ############################################################################################### for MOD
    # Class Function to create a temp sheet named '~-update' in order to update its content
    def Make_A_Same_Sheet_4Update(self, sheet_NameU, excel_FileNameU):
        update_Sheet_NameU = str(sheet_NameU).strip().replace("/", ".").lower() + '-update'
        wb = load_workbook(excel_FileNameU)
        a_IndexU = 0
        temp_CounterU = 0
        for sheet in wb.worksheets:
            if sheet == sheet_NameU:
                print("temp_CounterU: ", temp_CounterU)
                a_IndexU = temp_CounterU
                break
            temp_CounterU += 1
        wb.create_sheet(update_Sheet_NameU, a_IndexU + 1)

        sheet_Update = wb[update_Sheet_NameU]
        for i in self.Sch_1stRead:
            sheet_Update.append(i)
        wb.save(excel_FileNameU)
    
    ############################################################################################### for MOD
    ##### Calling 'Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule'
    # Class Function to read in the content of RCA's previously made schedule as 2nd for comparison
    def Read_A_Sheet_2nd(self, sheet_Name2):
        temp_Sheet_Name2 = str(sheet_Name2).strip().lower()
        if self.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(temp_Sheet_Name2, None)[0]:                                  
            temp_2Array2 = self.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(temp_Sheet_Name2, None)[1]
        else:
            RCA_Scheduler_Modifier.message_gen_final.changeText("Can\'t read a ", temp_Sheet_Name2, ' sheet') 
            RCA_Scheduler_Modifier.message_gen_final.changeColor('red')
        for inner_list2 in temp_2Array2:
            self.Sch_2ndRead_AfterModifying.append(inner_list2)

    ############################################################### for GEN
    # Class function to accumulate RCA's previous route work record
    def Accumulate_RCA_Previous_Routes(self, a_RTE_Schedule):
        rsA = re.compile('R\d{1,3}')                                                            # Regular Expression ex) R45
        for ii8 in range(0, len(self.RCA_Name)):                                                # Iteration with RCA names
            for jj8 in range(0, len(a_RTE_Schedule)):

                flag_Name_Matched = FALSE                                                       # Flag: RCA Name is matched with the name (the 1st item in the list of a RCA Schedule)
                if self.RCA_Name[ii8] == a_RTE_Schedule[jj8][0]:
                    flag_Name_Matched = TRUE
                    for kk8 in range(1, len(a_RTE_Schedule[jj8])):
                        if rsA.match(a_RTE_Schedule[jj8][kk8]) is not None:                                 # ^[DE][0-9]{2,3}$
                            if a_RTE_Schedule[jj8][kk8] in self.RCA_PreviousRoute_Info[ii8].keys():        # If a key exists, increase its value by 1
                                self.RCA_PreviousRoute_Info[ii8][a_RTE_Schedule[jj8][kk8]] += 1
                            else:                                                                           # If not existed, add it as a new key with 1 as its value
                                self.RCA_PreviousRoute_Info[ii8][a_RTE_Schedule[jj8][kk8]] = 1
                if flag_Name_Matched:
                    break

    ############################################################################# for GEN
    ##### Calling 'Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule' & 'Accumulate_RCA_Previous_Routes' & 'sort_Dic_ByValue'
    # Class function to read in multiple sheets in order to create a new schedule
    def Read_Multi_Sheets(self):
        list_SheetName_4MultiReading = self.get_Wb().sheetnames
        ## Sort out sheets with its name '*-leftover' if any
        regex = re.compile(r'\bleftover\b')
        filtered_list_SheetName_4MultiReading = [i for i in list_SheetName_4MultiReading if not regex.search(i)]
        
        # Add empty Dic into each RCA for collecting previous route working records
        for m2 in range(0, len(self.RCA_Name)):
            self.RCA_PreviousRoute_Info.append({})

        counter_Iter = 0
        # take only first 20 legitimate RCA schedule sheets
        for m3 in range(0, len(filtered_list_SheetName_4MultiReading)):
            if ('ww' in str(filtered_list_SheetName_4MultiReading[m3]).strip().lower() 
                and self.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(filtered_list_SheetName_4MultiReading[m3], None)[0] == TRUE):
                self.Accumulate_RCA_Previous_Routes(self.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(filtered_list_SheetName_4MultiReading[m3], None)[1])
                counter_Iter += 1

            if counter_Iter == 20:
                break

        # Reorder Dic by values and replaces old Dic
        for jk in range(0, len(self.RCA_Name)):
            reorder_Dic = {}
            reorder_Dic = self.sort_Dic_ByValue(self.RCA_PreviousRoute_Info[jk])
            del self.RCA_PreviousRoute_Info[jk]
            self.RCA_PreviousRoute_Info.insert(jk, reorder_Dic)
    

    #######################################  SMS Texting  #########################################

    ###################################################### for GEN & MOD
    ##### Calling 'create_Message_RCA_Sch' & 'create_Message_Updated_Sch'
    # Class Function to text a new schedule and any update
    def Texting_New_Sch_Or_Updated(self, mode03, sheet_Name03, sch_Value03):
        account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
        auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
        temp_account_phone = os.environ.get('TWILIO_ACCOUNT_PHONE')
        account_phone = '+1' + temp_account_phone.strip().replace(" ", "").replace("(", "").replace(")", "").replace("-", "")
        client = Client(account_sid, auth_token)

        message_SMS = []
        message_SMS_Diff = []

        if mode03 == 'gen':
            message_SMS = self.create_Message_RCA_Sch(sheet_Name03, sch_Value03)
            temp_Counter_GEN_Texting = 0
            for n04 in range(0, len(message_SMS)):
                if message_SMS[n04].get('Phone') is not None:
                    message = client.messages.create(
                                    body=message_SMS[n04].get('Body'), 
                                    from_= account_phone,
                                    to=message_SMS[n04].get('Phone')
                                )
                    message.sid
                temp_Counter_GEN_Texting += 1
            if temp_Counter_GEN_Texting == len(message_SMS) and self.Supervisor.get('Phone') is not None:
                message = client.messages.create(
                    body='RCA Scheduler of ' + sheet_Name03 + ' is successfully texted!', 
                    from_= account_phone,
                    to=self.Supervisor.get('Phone')
                      )
                message.sid

        if mode03 == 'mod':
            message_SMS_Diff = self.create_Message_Updated_Sch(sheet_Name03)
            temp_Counter_MOD_Texting = 0
            for n07 in range(0, len(message_SMS_Diff)):
                if message_SMS_Diff[n07].get('Phone') is not None:
                    message = client.messages.create(
                                    body=message_SMS_Diff[n07].get('Body'), 
                                    from_= account_phone,
                                    to=message_SMS_Diff[n07].get('Phone')
                                )
                    message.sid
                temp_Counter_MOD_Texting += 1
            if temp_Counter_MOD_Texting == len(message_SMS_Diff) and self.Supervisor.get('Phone') is not None:
                message = client.messages.create(
                    body='Updated RCA Schedules of ' + sheet_Name03 + ' are successfully texted!', 
                    from_= account_phone,
                    to=self.Supervisor.get('Phone')
                      )
                message.sid

            self.Empty_A_Variable_Sch_1stRead()
            self.Empty_A_Variable_Sch_2ndRead_AfterModifying()

    
    #################################################### for GEN or Mod
    # Class Function to find a route's station
    def find_Station_4Route(self, routeF):
        rsA = re.compile('R\d{1,3}')
        if rsA.match(routeF) is not None:                                 # ex)R45
            indexF = self.Route_Name.index(routeF)
            if self.Route_Value[indexF].get('Station') is not None:
                return '(' + self.Route_Value[indexF].get('Station') + ')'
            else:
                return ''
        else:
            return ''
        

    #################################################### for GEN
    # Class Function to create text message for each RCA
    def create_Message_RCA_Sch(self, sheet_Name04, sch_Value04):
        message_All_Rca = []
        temp_Sheet_Name04 = sheet_Name04.split('(')
        temp_Sheet_Name05 = '(' + temp_Sheet_Name04[1]
        for i4 in range(0, len(self.RCA_Name)):
            each_RCA_Msg = {}
            if self.RCA_Phone[i4] is not None:                                      # In case that phone number exists
                sms_Phone01 = str(self.RCA_Phone[i4]).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                each_RCA_Msg['Phone'] = '+1' + sms_Phone01
                each_RCA_Msg['Body'] = ('Hello, ' + self.RCA_Name[i4] + ';\n' + 'Your schedule of the week - ' + temp_Sheet_Name05 + '\n\n' 
                                            + 'Sat: ' + sch_Value04[i4][1] + self.find_Station_4Route(sch_Value04[i4][1]) + '\n'
                                            + 'Mon: ' + sch_Value04[i4][2] + self.find_Station_4Route(sch_Value04[i4][2]) + '\n'
                                            + 'Tue: ' + sch_Value04[i4][3] + self.find_Station_4Route(sch_Value04[i4][3]) + '\n'
                                            + 'Wed: ' + sch_Value04[i4][4] + self.find_Station_4Route(sch_Value04[i4][4]) + '\n'
                                            + 'Thu: ' + sch_Value04[i4][5] + self.find_Station_4Route(sch_Value04[i4][5]) + '\n'
                                            + 'Fri: ' + sch_Value04[i4][6] + self.find_Station_4Route(sch_Value04[i4][6]) + '\n'
                                            + '\nThanks')
                message_All_Rca.append(each_RCA_Msg)
            else:                                                                    # In case that phone number not exists
                message_All_Rca.append({'Phone': None, 'Body': None})
        return message_All_Rca
            

    ##################################################################### for GEN
    # Class Function to create text message for only updated RCS schedule
    def create_Message_Updated_Sch(self, sheet_Name06):
        message_Updated_Diff = []
        diff_2Schedules = []
        temp_Sheet_Name06 = sheet_Name06.split('(')
        temp_Sheet_Name07 = '(' + temp_Sheet_Name06[1]
        counter_Out = 0
        
        # Find difference between two schedules
        for item0 in self.Sch_1stRead:
            counter_In = 0
            flag_Diff = FALSE
            temp_Diff = {}
            if self.RCA_Phone[counter_Out] is not None:
                temp_Diff['Phone'] = '+1' + self.RCA_Phone[counter_Out].replace('-', '').replace('(', '').replace(')', '')
                temp_Diff['Name'] = self.RCA_Name[counter_Out]
                for item1 in item0:
                    if counter_In>=1 and counter_In<=6:
                        if item1 != self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]:
                            flag_Diff = TRUE
                            if counter_In == 1:
                                temp_Diff['Sat'] = self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]
                            elif counter_In == 2:
                                temp_Diff['Mon'] = self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]
                            elif counter_In == 3:
                                temp_Diff['Tue'] = self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]
                            elif counter_In == 4:
                                temp_Diff['Wed'] = self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]
                            elif counter_In == 5:
                                temp_Diff['Thu'] = self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]
                            elif counter_In == 6:
                                temp_Diff['Fri'] = self.Sch_2ndRead_AfterModifying[counter_Out][counter_In]

                    counter_In = counter_In + 1

                if flag_Diff:
                    diff_2Schedules.append(temp_Diff)
            else:                                                                    # In case that phone number not exists
                diff_2Schedules.append({'Phone': None})

            counter_Out = counter_Out + 1

        # Create Message based on difference found between two schedules
        for item in diff_2Schedules:
            each_Diff_Sch00 = {}
            if item['Phone'] is not None:
                each_Diff_Sch00['Phone'] = item['Phone']

                body_Part = ''
                body_Part = 'Hello, ' + item['Name'] + ';\n' + 'Your updated schedule of the week - ' + temp_Sheet_Name07 + '\n\n'
                for index, key in enumerate(item):
                    if index>=2:
                        body_Part = body_Part + key + ': ' + item.get(key) + self.find_Station_4Route(item.get(key))+ '\n'
                body_Part = body_Part + '\nThanks'
                each_Diff_Sch00['Body'] = body_Part    

                message_Updated_Diff.append(each_Diff_Sch00)
        
        return message_Updated_Diff


    #######################################  Creating RCA Scheduler  #######################################

    ############################################ for GEN
    ##### Calling 'Check_Day_2Num', 'Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2', & 'Assign_RTE_Based_On_Route_Value'
    # Class Function to create a schedule of RCA
    def Create_Scheduler(self, offDay_list0):

        # Make a list of busy days
        busy_Days = ['mon']
        for item in offDay_list0:
            if item == 'mon':
                busy_Days.remove('mon')
                busy_Days.append('tue')
            elif item == 'tue':
                if 'tue' in busy_Days:
                    busy_Days.remove('tue')
                busy_Days.append('wed')
            elif item == 'wed':
                if 'wed' in busy_Days:
                    busy_Days.remove('wed')
                busy_Days.append('thu')
            elif item == 'thu':
                if 'thu' in busy_Days:
                    busy_Days.remove('thu')
                busy_Days.append('fri')
            elif item == 'fri':
                if 'fri' in busy_Days:
                    busy_Days.remove('fri')
                busy_Days.append('sat')
        common_Item0 = set(offDay_list0) & set(busy_Days)
        if common_Item0 in offDay_list0 and common_Item0 in busy_Days:
            busy_Days.remove(common_Item0)

        # 
        ################# 1st round of creating schedule array by considering holiday, off-days (Lv), 
        #                                           training days, assigned off-days (L), assigned route, and primary route as this order
        for i in range(0, len(self.RCA_Name)):
            each_Sch = [None, None, None, None, None, None, None]

            # 1. Check Off Day
            if offDay_list0:
                for j in range(len(offDay_list0)):
                    if each_Sch[self.Check_Day_2Num(str(offDay_list0[j]).lower())] is None:
                        each_Sch[self.Check_Day_2Num(str(offDay_list0[j]).lower())] = 'Off'
            
            # 2. Check Lv
            if self.RCA_Value[i].get('Lv') is not None:
                Lv_list = str(self.RCA_Value[i].get('Lv')).replace(" ", '').split(',')
                for j1 in range(0, len(Lv_list)):
                    if each_Sch[self.Check_Day_2Num(str(Lv_list[j1]).lower())] is None:
                        each_Sch[self.Check_Day_2Num(str(Lv_list[j1]).lower())] = 'Lv'
            

            # 3, Check Traing days
            if self.RCA_Value[i].get('Training Days') is not None:
                training_list = str(self.RCA_Value[i].get('Training Days')).replace(" ", '').split(',')
                for j2 in range(0, len(training_list)):
                    if each_Sch[self.Check_Day_2Num(str(training_list[j2]).lower())] is None:
                        each_Sch[self.Check_Day_2Num(str(training_list[j2]).lower())] = 'Tr'

            # 4. Check L 
            if self.RCA_Value[i].get('L') is not None:                                  # Only one "L" value - (not conflicting with Lv, no L if Lv exists)

                flag_SameDay_Conflict = FALSE
                temp_Keys = ''
                # Check if conflicted with 'Lv', 'Off', and 'Tr'
                if 'Lv' in each_Sch:                                                # when 'Lv' existed
                    if self.Check_Day_2Num(str(self.RCA_Value[i].get('L')).lower()) == each_Sch.index('Lv'):
                        flag_SameDay_Conflict = TRUE                                # If true, assigning 'L' will be done next round
                        each_Sch[6] = 'L-Lv'
                if 'Tr' in each_Sch:                                                # when 'Tr' existed
                    if self.Check_Day_2Num(str(self.RCA_Value[i].get('L')).lower()) == each_Sch.index('Tr'):
                        flag_SameDay_Conflict = TRUE
                        each_Sch[6] = 'L-Tr'
                if 'Off' in each_Sch:                                                # when 'Off' existed
                    if self.Check_Day_2Num(str(self.RCA_Value[i].get('L')).lower()) == each_Sch.index('Off'):
                        flag_SameDay_Conflict = TRUE
                        each_Sch[6] = 'L-Off'
                # Check with Primary R if there is any conflict. If so, assigning 'L' moves into next round with the orignal 'L' value ignored
                if self.RCA_Value[i].get('Primary R') is not None:                      # Check if there is 'Primary R' & check if there is 'Lv'
                    index0 = self.Route_Name.index(self.RCA_Value[i].get('Primary R'))
                    for key, value in self.Route_Value[index0].items():
                        if value == 'L':                                                # Only one 'L'
                            temp_Keys=key
                    if self.Check_Day_2Num(str(self.RCA_Value[i].get('L')).lower()) == int(self.Check_Day_2Num(temp_Keys.lower())):
                        flag_SameDay_Conflict = TRUE
                        each_Sch[6] = 'L-PR'

                if not flag_SameDay_Conflict:
                     if each_Sch[self.Check_Day_2Num(str(self.RCA_Value[i].get('L')).lower())] is None:
                        each_Sch[self.Check_Day_2Num(str(self.RCA_Value[i].get('L')).lower())] = 'Lf'

            # 5. Check Assigned R
            if self.RCA_Value[i].get('Assigned R') is not None:
                remove_RTE_Value = []
                for nA in range(0, 6):
                    if each_Sch[nA] is None:
                        each_Sch[nA] = self.RCA_Value[i].get('Assigned R')
                        remove_RTE_Value.append(self.Check_Num_2Day(nA))

                # Come to Route_Value and then clear off corresponding 'Lv' or 'L'
                index_AR = self.Route_Name.index(self.RCA_Value[i].get('Assigned R'))
                for item in remove_RTE_Value:
                    self.Route_Value[index_AR][item] = None

            # 6. Check Primary R
            if self.RCA_Value[i].get('Primary R') is not None:

                index_PR = self.Route_Name.index(self.RCA_Value[i].get('Primary R'))
                for keyPR, valuePR in self.Route_Value[index_PR].items():
                    if valuePR == 'L' or valuePR == 'Lv':                                                    # Each route has only one 'L' and possibly multi 'Lv'
                        if each_Sch[self.Check_Day_2Num(keyPR.lower())] is None:
                            each_Sch[self.Check_Day_2Num(keyPR.lower())] = self.RCA_Value[i].get('Primary R')
                            # Come to Route_Value and then clear off corresponding 'L'
                            self.Route_Value[index_PR][keyPR] = None

    
            # Append each RCA's list into a whole 2 dimension list
            self.New_Sch.append(each_Sch)

        # Specially Handle Off-days for self.New_Sch
        if offDay_list0:
            for ii6 in range(len(offDay_list0)):
                for jj6 in range(0, len(self.Route_Name)):
                    # Come to Route_Value and then clear off corresponding 'L' or 'Lv'
                    if (self.Route_Value[jj6][str(offDay_list0[ii6]).lower().capitalize()] == 'L' 
                            or self.Route_Value[jj6][str(offDay_list0[ii6]).lower().capitalize()] == 'Lv'):
                        self.Route_Value[jj6][str(offDay_list0[ii6]).lower().capitalize()] = None


        #
        ## 2nd round of creating Schedule array by completing assigned off-days (L)
        #
        
        # 2-1. Assign 'L' if any RCA does not have 'L'
        if self.Mode_One_OffDay_Req == 'Yes':               # If one off-day is required
            for i2 in range(0, len(self.RCA_Name)):
                flag_L = FALSE
                if 'L' in self.New_Sch[i2] or 'Lf' in self.New_Sch[i2]:
                    flag_L = TRUE
                if not flag_L and None in self.New_Sch[i2]:                              
                    flag_Random_Num_Success = FALSE
                    while not flag_Random_Num_Success:
                        random_Num_4L = random.choices([0, 1, 2, 3, 4, 5], weights=(10, 20, 60, 60, 60, 20), k=1)
                        if self.New_Sch[i2][int(random_Num_4L[0])] is None:
                            self.New_Sch[i2][int(random_Num_4L[0])] = 'L'
                            flag_Random_Num_Success = TRUE


        # 
        ## 2-2. Consider 1st Secondary R, New Route Insertion Rate, Busy days, Station, and Seniority simultaneously
        #               Seniority - 4 Groups (5 days, 4 days, 3 days, 2 days)
        #               New RTE Insertion Rate -    0% (default): try to not insert any new RTE
        #                                           1% - 25%: try to add new RTE for only any one day by conisdering only their working station
        #                                           25% - 50%: try to add new RTE for two days at Max by conisdering only their working station
        #                                           51% - 75%: try to add new RTE for three days at Max by conisdering only their working station
        #                                           76% - 100%: try to add new RTE for four days at Max by conisdering only their working station
        divide_Num = len(self.RCA_Name) // 4
        for i3 in range(0, len(self.RCA_Name)):
            list_NONE = []
            rca_P_Station = self.RCA_Value[i3]['Preferred Station']
            for jj3 in range(0, 6):
                if self.New_Sch[i3][jj3] is None:
                    list_NONE.append(jj3)

            # Default Value of New RTE Insertion Rate (No day for New RTE)
            if self.Get_New_Route_Insertion_Rate() == 0:
                if i3 >= 0 and i3 < divide_Num:                                             # 1st Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 0, rca_P_Station, busy_Days)
                elif divide_Num >= 0 and i3 < (divide_Num * 2):                             # 2nd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 0, rca_P_Station, busy_Days)
                elif (divide_Num * 2) >= 0 and i3 < (divide_Num * 3):                       # 3rd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 1, rca_P_Station, busy_Days)
                else:                                                                      # 4th Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 2, rca_P_Station, busy_Days)

            # New RTE Insertion Rate 1 < RTE < 25
            elif self.Get_New_Route_Insertion_Rate() >= 1 and self.Get_New_Route_Insertion_Rate() <= 25:
                if i3 >= 0 and i3 < divide_Num:                                         # 1st Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 1, rca_P_Station, busy_Days)
                elif divide_Num >= 0 and i3 < (divide_Num * 2):                             # 2nd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 1, rca_P_Station, busy_Days)
                elif (divide_Num * 2) >= 0 and i3 < (divide_Num * 3):                    # 3rd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 2, rca_P_Station, busy_Days)
                else:                                                                      # 4th Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 3, rca_P_Station, busy_Days)
            # New RTE Insertion Rate 26 < RTE < 50
            elif self.Get_New_Route_Insertion_Rate() >= 26 and self.Get_New_Route_Insertion_Rate() <= 50:
                if i3 >= 0 and i3 < divide_Num:                                         # 1st Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 2, rca_P_Station, busy_Days)
                elif divide_Num >= 0 and i3 < (divide_Num * 2):                             # 2nd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 2, rca_P_Station, busy_Days)
                elif (divide_Num * 2) >= 0 and i3 < (divide_Num * 3):                    # 3rd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 3, rca_P_Station, busy_Days)
                else:                                                                      # 4th Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 4, rca_P_Station, busy_Days)
            # New RTE Insertion Rate 51 < RTE < 75
            elif self.Get_New_Route_Insertion_Rate() >= 51 and self.Get_New_Route_Insertion_Rate() <= 75:
                if i3 >= 0 and i3 < divide_Num:                                         # 1st Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 3, rca_P_Station, busy_Days)
                elif divide_Num >= 0 and i3 < (divide_Num * 2):                             # 2nd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 3, rca_P_Station, busy_Days)
                elif (divide_Num * 2) >= 0 and i3 < (divide_Num * 3):                    # 3rd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 4, rca_P_Station, busy_Days)
                else:                                                                      # 4th Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 5, rca_P_Station, busy_Days)
            # New RTE Insertion Rate 76 < RTE < 100
            elif self.Get_New_Route_Insertion_Rate() >= 76 and self.Get_New_Route_Insertion_Rate() <= 100:
                if i3 >= 0 and i3 < divide_Num:                                         # 1st Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 4, rca_P_Station, busy_Days)
                elif divide_Num >= 0 and i3 < (divide_Num * 2):                             # 2nd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 4, rca_P_Station, busy_Days)
                elif (divide_Num * 2) >= 0 and i3 < (divide_Num * 3):                    # 3rd Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 5, rca_P_Station, busy_Days)
                else:                                                                      # 4th Group
                    self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(i3, list_NONE, 6, rca_P_Station, busy_Days)

        # 
        ## 3rd round of creating Schedule array by completing assigned off-days using only Route value randomly
        #
        for i44 in range(0, len(self.RCA_Name)):
            list_NONE5 = []
            for jj4 in range(0, 6):
                if self.New_Sch[i44][jj4] is None:
                    list_NONE5.append(jj4)
            
            temp_Counter4 = 0
            while len(list_NONE5) > 0:
                an_Index_Random4 = random.choices(list_NONE5, k=1)
                self.Assign_RTE_Based_On_Route_Value(i44, an_Index_Random4[0], list_NONE5, None)
                if temp_Counter4 == 6:
                    break
                temp_Counter4 += 1

        # 
        ## To show RTEs whose open days are not filled - 2 diemensional lists of leftover
        #
        Days = ['Sat', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        for i5 in range(0, len(self.Route_Name)):
            each_List_RTE_LeftOver = []
            counter_LeftOver = 0
            for key5, value5 in self.Route_Value[i5].items():
                if key5 in Days and (value5 == 'L' or value5 == 'Lv'):
                    if counter_LeftOver == 0:
                        each_List_RTE_LeftOver.append(self.Route_Name[i5])
                        each_List_RTE_LeftOver.append(key5)
                    else:
                        each_List_RTE_LeftOver.append(key5)
                    counter_LeftOver += 1
            if len(each_List_RTE_LeftOver) != 0:
              self.Route_Value_LeftOver.append(each_List_RTE_LeftOver)

        # 
        ## 4th round to adjust the generated schedule based on leftover once again
        #

        for i6 in range(0, len(self.Route_Value_LeftOver)):
            for j6 in range(1, len(self.Route_Value_LeftOver[i6])):
                # find right RCA whose days are open and self.Route_Value_LeftOver[i6][j6] is 'L'
                index_L6 = self.Check_Day_2Num(str(self.Route_Value_LeftOver[i6][j6]).lower())
                for ii6 in range(0, len(self.New_Sch)):
                    if self.New_Sch[ii6][index_L6] == 'L' and None in self.New_Sch[ii6][:-1]:
                        list_Index_None6 = [i for i, x in enumerate(self.New_Sch[ii6][:-1]) if x == None]
                        random_Index6 = random.choices(list_Index_None6, k=1)
                        self.New_Sch[ii6][random_Index6[0]] = 'L'
                        self.New_Sch[ii6][index_L6] = self.Route_Value_LeftOver[i6][0]                  # a RTE is assigned by replacing 'L'
                        # Apply a change into Route_Value
                        index_RTE6 = self.Route_Name.index(self.Route_Value_LeftOver[i6][0])
                        self.Route_Value[index_RTE6][self.Route_Value_LeftOver[i6][j6]] = None
                        self.Route_Value_LeftOver[i6].pop(j6)
                        # flag_Found_Replacement = TRUE
                        break
                
        list_No_leftOver = []
        for ih6 in range(0, len(self.Route_Value_LeftOver)):
            if len(self.Route_Value_LeftOver[ih6]) == 1:
                list_No_leftOver.append(ih6)

        for ihh6 in range(len(list_No_leftOver)-1, -1, -1):
            self.Route_Value_LeftOver.pop(ihh6)

        # 
        ## 5th round of creating Schedule array by adding report time into blank slot
        #
        for i6 in range(0, len(self.RCA_Name)):
            report_Time = self.RCA_Value[i6]['Report Hour']
            for jj6 in range(0, 6):
                if self.New_Sch[i6][jj6] is None:
                    self.New_Sch[i6][jj6] = report_Time
                if self.New_Sch[i6][jj6] == 'Lf':
                    self.New_Sch[i6][jj6] = 'L'


    ########################################################### for GEN
    ##### Calling 'Assign_RTE_2Open_Days_Based_On_Various_Ways_Level1'
    # Class Function to find rigt Route for open days - level 2
    def Assign_RTE_2Open_Days_Based_On_Various_Ways_Level2(self, RCA_counter77, list_NONE77, limit77, station77, busy_Days77):
         # For Busy Days first
        for busy in busy_Days77:
            # handle each busy day one by one
            temp_NONE_List01 = self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level1(RCA_counter77, self.Check_Day_2Num(busy), list_NONE77, limit77, station77)
            if len(temp_NONE_List01) != len(list_NONE77):
                list_NONE77.clear()
                for item88 in temp_NONE_List01:
                    list_NONE77.append(item88)
                    
        # Randomly Pick any open days and apply various ways to asssign RTE into them
        temp_Counter = 0
        while len(list_NONE77) > limit77:
            a_Index_4Random_Day = random.choices(list_NONE77, k=1)
            temp_NONE_List02 = self.Assign_RTE_2Open_Days_Based_On_Various_Ways_Level1(RCA_counter77, a_Index_4Random_Day[0], list_NONE77, limit77, station77)
            if len(temp_NONE_List02) == 0:
                    break
            if len(temp_NONE_List02) != len(list_NONE77):
                list_NONE77.clear()
                for item77 in temp_NONE_List02:
                    list_NONE77.append(item77)
            temp_Counter += 1
            if temp_Counter == 6:
                break

    ########################################################### for GEN
    ##### Calling 'Assign_RTE_Based_On_Its_Previous_Record'
    # Class Function to find rigt Route for open days - level 1
    def Assign_RTE_2Open_Days_Based_On_Various_Ways_Level1(self, RCA_counter99, index_PR99, list_NONE99, limit99, station99):
        temp_List_NONE99 = []
        for item99 in list_NONE99:
            temp_List_NONE99.append(item99)

        if index_PR99 in temp_List_NONE99 and len(temp_List_NONE99) > limit99:
            temp_Size_Before0 = len(temp_List_NONE99)
            self.Assign_RTE_Based_On_Its_Previous_Record(RCA_counter99, index_PR99, temp_List_NONE99)
            if temp_Size_Before0 == len(temp_List_NONE99):
                temp_Size_Before1 = len(temp_List_NONE99)
                self.Assign_RTE_Based_On_Secondary_R1(RCA_counter99, index_PR99, temp_List_NONE99)
                if temp_Size_Before1 == len(temp_List_NONE99):
                    temp_Size_Before2 = len(temp_List_NONE99)
                    self.Assign_RTE_Based_On_Secondary_R2(RCA_counter99, index_PR99, temp_List_NONE99)
                    if temp_Size_Before2 == len(temp_List_NONE99):
                        self.Assign_RTE_Based_On_Route_Value(RCA_counter99, index_PR99, temp_List_NONE99, station99)
        return temp_List_NONE99

    ################################################################################ for GEN
    ##### Calling 'Check_Day_2Num'
    # Class Function to assign RTE based on its previous TCA's RTE assignment record
    def Assign_RTE_Based_On_Its_Previous_Record(self, RCA_counter, index_PR11, list_NONE11):

        ############# Presetting                            
        flag_FOUND_Available_RTE_From_PreRecord = FALSE
        #############
        if len(self.RCA_PreviousRoute_Info) != 0:
            for key11, value11 in self.RCA_PreviousRoute_Info[RCA_counter].items():
                index_11R = self.Route_Name.index(key11)

                for key12, value12 in self.Route_Value[index_11R].items():
                    if value12 == 'L'  or  value12 == 'Lv':
                        if self.Check_Day_2Num(str(key12).lower()) == index_PR11:
                            flag_FOUND_Available_RTE_From_PreRecord = TRUE
                            self.New_Sch[RCA_counter][index_PR11] = key11
                            # Clear off 'L' or 'Lv' from self.Route_Value
                            self.Route_Value[index_11R][key12] = None
                            list_NONE11.remove(index_PR11)
                            break
                if flag_FOUND_Available_RTE_From_PreRecord:
                    break
        # return list_NONE11
    
    ######################################################## for GEN
    ##### Calling 'Check_Day_2Num'
    # Class Function to assign RTE based on the secondary R1
    def Assign_RTE_Based_On_Secondary_R1(self, RCA_counter0, index_PR22, list_NONE22):

        # flag_FOUND_Available_RTE_From_SecondaryR1 = FALSE
        if self.RCA_Value[RCA_counter0]['1st Secondary R'] is not None:
            index_4SecondaryR1 = self.Route_Name.index(self.RCA_Value[RCA_counter0]['1st Secondary R'])

            for keySR1, valueSR1 in self.Route_Value[index_4SecondaryR1].items():
                if valueSR1 == 'L' or valueSR1 == 'Lv':
                    if self.Check_Day_2Num(str(keySR1).lower()) == index_PR22:
                        # flag_FOUND_Available_RTE_From_SecondaryR1 = TRUE
                        self.New_Sch[RCA_counter0][index_PR22] = self.Route_Name[index_4SecondaryR1]
                        # Clear off 'L' or 'Lv' from self.Route_Value
                        self.Route_Value[index_4SecondaryR1][keySR1] = None
                        list_NONE22.remove(index_PR22)
                        break
        # return list_NONE22

    ######################################################## for GEN
    ##### Calling 'Check_Day_2Num'
    # Class Function to assign RTE based on the secondary R2
    def Assign_RTE_Based_On_Secondary_R2(self, RCA_counter1, index_PR33, list_NONE33):

        # flag_FOUND_Available_RTE_From_SecondaryR2 = FALSE
        if self.RCA_Value[RCA_counter1]['2nd Secondary R'] is not None:
            index_4SecondaryR2 = self.Route_Name.index(self.RCA_Value[RCA_counter1]['2nd Secondary R'])

            for keySR2, valueSR2 in self.Route_Value[index_4SecondaryR2].items():
                if valueSR2 == 'L' or valueSR2 == 'Lv':
                    if self.Check_Day_2Num(str(keySR2).lower()) == index_PR33:
                        # flag_FOUND_Available_RTE_From_SecondaryR2 = TRUE
                        self.New_Sch[RCA_counter1][index_PR33] = self.Route_Name[index_4SecondaryR2]
                        # Clear off 'L' or 'Lv' from self.Route_Value
                        self.Route_Value[index_4SecondaryR2][keySR2] = None
                        list_NONE33.remove(index_PR33)
                        break
        # return list_NONE33

    ################################################################################################# for GEN
    ##### Calling 'Check_Day_2Num'
    # Class Function to pick a RTE by searching thru from top of Route Value conisdering same station
    def Assign_RTE_Based_On_Route_Value(self, RCA_counter2, index_PR44, list_NONE44, rca_P_Station44):

        list_Available_RTE_Index = []
        list_All_Num = []
        for iter in range(0, len(self.Route_Value)):
            list_All_Num.append(iter)

        for ttE in range(0, len(self.Route_Value)):
            random_One_Num = random.choice(list_All_Num)
            for keyE, valueE in self.Route_Value[random_One_Num].items():
                if keyE == self.Check_Num_2Day(index_PR44).capitalize() and (valueE == 'L' or valueE == 'Lv'):
                    list_Available_RTE_Index.append({'Route': self.Route_Name[random_One_Num], 'Station': self.Route_Value[random_One_Num]['Station']})
            list_All_Num.remove(random_One_Num)

        for ttF in range(0, len(list_Available_RTE_Index)):
            if rca_P_Station44 is None:                                                   # In case that it does not have any preferred station
                self.New_Sch[RCA_counter2][index_PR44] = list_Available_RTE_Index[ttF]['Route']
                index_F = self.Route_Name.index(list_Available_RTE_Index[ttF]['Route'])
                self.Route_Value[index_F][self.Check_Num_2Day(index_PR44).capitalize()] = None
                list_NONE44.remove(index_PR44)
                break
            else:
                if rca_P_Station44 == list_Available_RTE_Index[ttF]['Station']:                          # Check index_PR44
                    self.New_Sch[RCA_counter2][index_PR44] = list_Available_RTE_Index[ttF]['Route']
                    index_F = self.Route_Name.index(list_Available_RTE_Index[ttF]['Route'])
                    self.Route_Value[index_F][self.Check_Num_2Day(index_PR44).capitalize()] = None
                    list_NONE44.remove(index_PR44)
                    break
        # return list_NONE44
    

    #######################################  Writing RCA Scheduler into a Excel #######################################

    ########################################################## for GEN
    # Class Function to Write a Generted Schedule into a Sheet
    def Write_Generated_Sch_2Sheet(self, sheet_Name0, excel_FileName0):

        # Make DataFrame from the List of new Schedule
        for i in range(0, len(self.New_Sch)):
            self.New_Sch[i].insert(0, self.RCA_Name[i])

        # Create a sheet at 3rd place and write a new schedule into it
        new_Sheet_Name0 = str(sheet_Name0).strip().replace('/', '.')
        wb = load_workbook(excel_FileName0)
        wb.create_sheet(new_Sheet_Name0, 2)
        sheet_new = wb[new_Sheet_Name0]
        for i in self.New_Sch:
            sheet_new.append(i)
        wb.save(excel_FileName0)

        # Create a sheet of '-leftover' at 4th place and write self.Route_Value_LeftOver into it
        if len(self.Route_Value_LeftOver):
            new_Sheet_Name_Left_Over = new_Sheet_Name0 + '-leftover'
            wb_2nd = load_workbook(excel_FileName0)
            wb_2nd.create_sheet(new_Sheet_Name_Left_Over, 3)
            sheet_new_leftover=wb_2nd[new_Sheet_Name_Left_Over]
            for j in self.Route_Value_LeftOver:
                sheet_new_leftover.append(j)
            wb_2nd.save(excel_FileName0)

     ############################################################## for GEN
    # Class Function to set the size of cells for new RCA Schedule
    def Adjust_Size_Of_Cells(self, sheet_Name1, excel_FileName1):

        book1 = load_workbook(excel_FileName1)
        new_Sheet_Name0 = sheet_Name1.split('(')
        new_Sheet_Name1 = new_Sheet_Name0[1].replace('/', '.')
        new_Sheet_Name2 = new_Sheet_Name0[0] + '(' + new_Sheet_Name1
        sheet_Int = book1[new_Sheet_Name2]

        # Make necessary changes by creating two first rows and set their row size
        sheet_Int.insert_rows(1)
        sheet_Int.insert_rows(1)
        sheet_Int.row_dimensions[2].height = 13
        
        sheet_Int.row_dimensions[1].height = 30
        sheet_Int.column_dimensions['A'].width = 18

        # Create new columns
        for nm03 in range(8, 1,-1):
            sheet_Int.insert_cols(nm03)
        sheet_Int.delete_cols(2)

        ######################## Write to cells
        sheet_Int.cell(row=1, column=1).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[0]
        sheet_Int.cell(row=1, column=2).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[1]
        sheet_Int.cell(row=1, column=4).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[2]
        sheet_Int.cell(row=1, column=6).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[3]
        sheet_Int.cell(row=1, column=8).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[4]
        sheet_Int.cell(row=1, column=10).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[5]
        sheet_Int.cell(row=1, column=12).value = self.Convert_Sheet_Name_2Week_Days(sheet_Name1)[6]
        sheet_Int.cell(row=1, column=14).value = 'Comments'

        for kk in range(2, 14, 2):
            sheet_Int.cell(row=2, column=kk).value = 'RTE'
        for kl in range(3, 15, 2):
            sheet_Int.cell(row=2, column=kl).value = 'Others'

        ####################### Set Column Width
        list_Aph_Route = ['B', 'D', 'F', 'H', 'J', 'L']
        for nn in list_Aph_Route:
            sheet_Int.column_dimensions[nn].width = 5
        #######################
        list_Aph_Space = ['C', 'E', 'G', 'I', 'K', 'M']
        for nn in list_Aph_Space:
            sheet_Int.column_dimensions[nn].width = self.Get_New_Cell_Column_Width() - 5
        sheet_Int.column_dimensions['N'].width = 14
        # ####################### Merge Cells
        sheet_Int.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)             # RCA NAme
        sheet_Int.merge_cells(start_row=1, end_row=2, start_column=14, end_column=14)             # Comments
        sheet_Int.merge_cells(start_row=1, end_row=1, start_column=2, end_column=3)
        sheet_Int.merge_cells(start_row=1, end_row=1, start_column=4, end_column=5)
        sheet_Int.merge_cells(start_row=1, end_row=1, start_column=6, end_column=7)
        sheet_Int.merge_cells(start_row=1, end_row=1, start_column=8, end_column=9)
        sheet_Int.merge_cells(start_row=1, end_row=1, start_column=10, end_column=11)
        sheet_Int.merge_cells(start_row=1, end_row=1, start_column=12, end_column=13)

        # Pick rows after first row
        for row_num in range(3, len(self.RCA_Name) + 3):
            sheet_Int.row_dimensions[row_num].height = 18
        
        # Pick rows after first row and then apply setting of cell format
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        colorFill_Title = PatternFill(start_color='99FFCC5F',
                   end_color='99FFCC5F',
                   fill_type='solid')
        colorFill_SubTitle = PatternFill(start_color='99FFCC5F',
                   end_color='99FFCC5F',
                   fill_type='solid')
        colorFill_Name = PatternFill(start_color='99FFFC65',
                   end_color='99FFFC65',
                   fill_type='solid')
        colorFill_Item = PatternFill(start_color='1AEBEBEB',
                   end_color='1AEBEBEB',
                   fill_type='solid')
        
        for row1 in sheet_Int.iter_rows():
            if row1[0].row == 1:
                for cell in row1:
                    if cell.column>=1 and cell.column<=14:
                        cell.font = Font(bold=True, size='12')
                        cell.border = thin_border
                        cell.fill = colorFill_Title
                        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    if cell.column==14:
                        break
            if row1[0].row == 2:
                for cell in row1:
                    if cell.column>=2 and cell.column<=13:
                        cell.font = Font(bold=True, size='11')
                        cell.border = thin_border
                        cell.fill = colorFill_Title
                        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    if cell.column==13:
                        break

            if row1[0].row>2:
                for cell in row1:
                    if cell.column==1:
                        cell.font = Font(bold=True, size='11')
                        cell.border = thin_border
                        cell.fill = colorFill_Name
                    else:
                        if (cell.column>=2 and cell.column<=13) and (cell.column % 2) == 0:
                            cell.border = thin_border
                            cell.font = Font(size='11')
                            cell.fill = colorFill_Item
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                        else:
                            cell.font = Font(size='11')
                            cell.border = thin_border
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    if cell.column==14:
                        break
            
            # save the file

        book1.save(excel_FileName1)


#
################################### Class of 'RCA_Scheduler_Modifier' ############################
#

################################################### Class 'RCA_Scheduler_Modifier'
# Class of GUI of RCA Schedule Generator & Modifier
class RCA_Scheduler_Modifier:

    ############################################## Class Function
    # __init__
    def __init__(self, master):

        self.myExcel = Excel()

        self.resultFlag = ["", "", "", FALSE, []]
        self.GUI(master)

    ############################################## Class Function
    # Class Function for GUI
    def GUI(self, masterR):

        # 1st Label in Generator
        self.label_gen0 = label_gui(masterR)
        self.label_gen0.setLabel('RCA Scheduler', 'Helvetica 11 bold', 8)
        self.label_gen0.setGridLabel(0, 0, 3)

        # Label for Open File in Generator
        self.label_gen_openfile = label_gui(masterR)
        self.label_gen_openfile.setLabel('Open an Excel File', 'Helvetica 9 bold', 8)
        # self.label_gen_openfile.setLabel('Open an Excel File', 'Helvetica 10 bold', 8)
        self.label_gen_openfile.setGridLabel(0, 1, 2)

        # Label to diplay an Opened File in Generator
        self.label_gen_displayopenfile = label_gui(masterR)
        self.label_gen_displayopenfile.setLabel0('File Path: ', 'Helvetica 9', 25, 5, LEFT, 7, 'blue')
        self.label_gen_displayopenfile.setGridLabel0('W', 0, 2, 2)

        # Button for Open File in Generator
        self.button_gen_openfile = button_gui(masterR)
        self.button_gen_openfile.setButton('1. Open an Excel File', 20, 'white', 5, 'blue', lambda:self.Click_OpenFile(1))
        self.button_gen_openfile.setGridButton(0, 3, 2, 0)

        # Entry for Name of Sheet in Generator
        self.entry_gen_forsheet = entry_gui(masterR)
        self.entry_gen_forsheet.setEntry('', 25, 3, 'light gray')
        self.entry_gen_forsheet.setGridEntry(2, 1, 1, 0)
        self.entry_gen_forsheet.deleteValueEntry()
        self.entry_gen_forsheet.setValueEntry("Enter Sheet's Name")
        # self.entry_gen_forsheet.bindClick()

        # Entry for Off Days in Generator
        self.entry_gen_off = entry_gui(masterR)
        self.entry_gen_off.setEntry('', 25, 3, 'light gray')
        self.entry_gen_off.setGridEntry(2, 2, 1, 0)
        self.entry_gen_off.deleteValueEntry()
        self.entry_gen_off.setValueEntry("Enter Off-Day(e.g.: Mon,Thu)")

        # Button for Schedule Generation in Generator
        self.button_gen0 = button_gui(masterR)
        self.button_gen0.setButton('2. Generate a Schedule', 20, 'gray', 5, 'yellow', lambda:self.Click_Sch_Generate())
        self.button_gen0.setGridButton(2, 3, 1, 0)

        # Message for Schedule Generation in Generator
        self.message_gen0 = message_gui(masterR)
        self.message_gen0.setMessage("", 'Helvetica 9', 'orange', 250, 2)
        # self.message_gen0.setMessage("Make changes on restrictions before generating", 'Helvetica 9', 'orange', 250, 2)
        self.message_gen0.setGridMessage(2, 4, 1)

        # Button for Text Schedule in Generator
        self.button_gen_textSch = button_gui(masterR)
        self.button_gen_textSch.setButton('3. Text a Schedule', 20, "gray", 5, "yellow", lambda:self.Click_Txt_Sch())
        self.button_gen_textSch.setGridButton(2, 5, 1, 0)

        # Message for Text Schedule in Generator
        self.message_gen_textSch = message_gui(masterR)
        self.message_gen_textSch.setMessage("Save any Change before Texting", 'Helvetica 9', 'purple', 700, 2)
        # self.message_gen_textSch.setMessage("Save any Change, Close, & Texting", 'Helvetica 9', 'purple', 250, 2)
        self.message_gen_textSch.setGridMessage(2, 6, 1)

        # Final Message for Generator
        self.message_gen_final = message_gui(masterR)
        self.message_gen_final.setMessage("", 'Helvetica 10', 'green', 1000, 6)
        self.message_gen_final.setGridMessage(0, 7, 3)

        # Border Line between Generator and Modifier
        for x in range(3):
            canvas=tk.Canvas(masterR, width='15',height='3',bg='black')                                           
            canvas.bind("<Configure>", self.Redraw_Line)
            canvas.create_line(0,0,0,0, tags=("line1",))
            canvas.grid(row=8,column=x,sticky='EW')
        for x in range(3):
            masterR.columnconfigure(x,weight=1)
            masterR.rowconfigure(8, pad=1)

        # ################### Modifier Part #####################
        # 1st Label in Modifier
        self.label_mod0 = label_gui(masterR)
        self.label_mod0.setLabel('RCA Schedule Updater', 'Helvetica 11 bold', 8)
        self.label_mod0.setGridLabel(0, 9, 3)

        # 1st Label in Modifier
        self.label_mod0 = label_gui(masterR)
        self.label_mod0.setLabel('(1. Read in Sheet\'s Contents -> Modify -> 2. Texting)', 'Helvetica 9 bold', 1)
        self.label_mod0.setGridLabel(0, 10, 3)

        # Label for Open File in Modifier
        self.label_mod_openfile = label_gui(masterR)
        self.label_mod_openfile.setLabel('Open an Excel File', 'Helvetica 9 bold', 8)
        self.label_mod_openfile.setGridLabel(0, 11, 2)

        # Label to diplay an Opened File in Modifier
        self.label_mod_displayopenfile = label_gui(masterR)
        self.label_mod_displayopenfile.setLabel0('File Path: ', 'Helvetica 9', 27, 5, LEFT, 5, 'blue')
        self.label_mod_displayopenfile.setGridLabel0('W', 0, 12, 2)

        # Button for Open File in Modifier
        self.button_mod_openfile = button_gui(masterR)
        self.button_mod_openfile.setButton('1-1. Open an Excel File', 20, 'white', 5, 'blue', lambda:self.Click_OpenFile(2))
        self.button_mod_openfile.setGridButton(0, 13, 2, 5)

        # Entry for Name of Sheet in Modifer
        self.entry_mod_forsheet = entry_gui(masterR)
        self.entry_mod_forsheet.setEntry('', 30, 3, 'light gray')
        self.entry_mod_forsheet.setGridEntry(0, 14, 2, 5)
        self.entry_mod_forsheet.deleteValueEntry()
        self.entry_mod_forsheet.setValueEntry("Enter Sheet's Name")
        # self.entry_mod_forsheet.bindClick()

        # Button for Schedule Modify in Modfifier
        self.button_mod0 = button_gui(masterR)
        self.button_mod0.setButton('1-2. Read in Sheet\'s Contents', 30, 'gray', 5, 'yellow', lambda:self.Click_Get_SheetName())
        self.button_mod0.setGridButton(0, 15, 2, 0)

        # Message for Schedule Modify in Modfifier
        self.message_mod0 = message_gui(masterR)
        self.message_mod0.setMessage("Stop Here!! -> Make changes in Excel, Save, & Close it -> Click 'Text Update'", 'Helvetica 9 bold', 'purple', 800, 2)
        self.message_mod0.setGridMessage(0, 16, 2)

        # Button for Text Update in Modfifier
        self.button_mod_textSch = button_gui(masterR)
        self.button_mod_textSch.setButton('2. Text Update', 20, "gray", 5, "yellow", lambda:self.Click_Txt_Update())
        self.button_mod_textSch.setGridButton(2, 12, 1, 0)

        # Button for SetUp
        self.button_setup = button_gui(masterR)
        self.button_setup.setButton('SET UP', 13, "white", 5, "orange", lambda: self.Click_Setup(masterR))
        # self.button_setup.setButton('Set Up', 15, "white", 5, "orange", lambda: self.Click_Setup(master))
        # self.button_setup.setGridButton(2, 15, 1, 0)
        self.button_setup.setGridButton(2, 15, 1, 7)

        # Final Message for Updater
        self.message_mod_final = message_gui(masterR)
        self.message_mod_final.setMessage("", 'Helvetica 10', 'green', 1000, 6)
        self.message_mod_final.setGridMessage(0, 17, 3)

        # Button for HELP
        self.button_help = button_gui(masterR)
        self.button_help.setButton('CLOSE', 13, "white", 5, "black", lambda:self.Click_Close_Main(masterR))
        self.button_help.setGridButton(2, 16, 1, 0)

    ############################################## Class Function
    # Class Function to redraw Line in GUI
    def Redraw_Line(self, event):
        width = event.width
        canvas = event.widget
        canvas.coords("line1", 0, 0, width, 5)

    ############################################## Class Function
    # Class Function to click to open an excel file
    def Click_OpenFile(self, flagSC):
        afile = self.OpenFile()
        if flagSC == 1:                          # GEN
            self.label_gen_displayopenfile.configureText('File Path: ' + afile)
            # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
            self.resultFlag[0] = "gen"
            self.resultFlag[1] = afile
            self.resultFlag[2] = ''
            self.resultFlag[3] = FALSE
            self.resultFlag[4] = []
        else:                                   # MOD
            self.label_mod_displayopenfile.configureText('File Path: ' + afile)
            # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
            self.resultFlag[0] = "mod"
            self.resultFlag[1] = afile
            self.resultFlag[2] = ''
            self.resultFlag[3] = FALSE

    #################################################### Class Function
    # Class Function to open an excel file and file name
    def OpenFile(self):
        self.file_path = askopenfilename(filetypes=[('Excel Files', '*.xlsx *.xlsm *.xltx *.xltm *.csv')])
        if self.file_path is not None:
            return self.file_path

    ######################################################## Class Function
    # Class Function to click to generate a new RCA schedule
    def Click_Sch_Generate(self):
        # check format of sheet's name
        rS = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\/\d{1,2}\/\d{1,2}-\d{1,2}\/\d{1,2}\/\d{1,2}\)')
        rSA = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\.\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}\.\d{1,2}\)')
        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
        if self.resultFlag[1] != '':                                    # Excel File Check
            if ((rS.match(str(self.entry_gen_forsheet.getText()).replace(" ","").replace("\n", "")) is not None 
                    or rSA.match(str(self.entry_gen_forsheet.getText()).replace(" ","").replace("\n", "")) is not None)
                        and self.Days_Between(str(self.entry_gen_forsheet.getText()).replace(" ","").replace("\n", ""))==6):      # Sheet Name Check
                # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                self.resultFlag[2] = str(self.entry_gen_forsheet.getText()).strip().replace(" ", "").lower()
                # check format of off days
                if self.entry_gen_off.getText() != '' and self.entry_gen_off.getText() != "Enter Off-Day(e.g.: Mon,Thu)":    # if not empty
                    offDayS0 = str(self.entry_gen_off.getText()).strip().replace(" ", "")
                    offDayS1 = offDayS0.split(',')
                    offDayS1 = [each_str.lower() for each_str in offDayS1]
                    if self.Check_Format_ListsOfOffDays(offDayS1) == TRUE:                      # Right Format of Off-days Check if any - # Success
                        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                        self.resultFlag[4] = offDayS1
                        self.resultFlag[3] = FALSE                                           

                        # ################## Start Working on Excel File
                        returned_Value_FromExcel = self.Working_With_Excel()

                        if returned_Value_FromExcel[0]:
                            temp_Sheet_Name = self.resultFlag[2].replace('/', '.')
                            temp_Msg = returned_Value_FromExcel[1] + ' - ' + temp_Sheet_Name
                            self.message_gen_final.changeText(temp_Msg)
                            self.message_gen_final.changeColor('green')
                        else:
                            self.message_gen_final.changeText(returned_Value_FromExcel[1])
                            self.message_gen_final.changeColor('red')
                        
                        
                    else:
                        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                        self.resultFlag[4] = []
                        self.message_gen_final.changeText('Generating a RCA Schedule Fails!! - Wrong Off-Days Format')
                        self.message_gen_final.changeColor('red')
                else:
                    # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                    self.resultFlag[4] = []
                    self.resultFlag[3] = FALSE

                    # ################## Start Working on Excel File
                    returned_Value_FromExcel = self.Working_With_Excel()

                    if returned_Value_FromExcel[0]:
                        temp_Sheet_Name = self.resultFlag[2].replace('/', '.')
                        temp_Msg = returned_Value_FromExcel[1] + ' - ' + temp_Sheet_Name
                        self.message_gen_final.changeText(temp_Msg)
                        self.message_gen_final.changeColor('green')
                    else:
                        self.message_gen_final.changeText(returned_Value_FromExcel[1])
                        self.message_gen_final.changeColor('red')
            else:
                self.message_gen_final.changeText('Generating a RCA Schedule Fails!! - Missing or Wrong Sheet\'s Name')
                self.message_gen_final.changeColor('red')
        else:                                                           # Fail - Wrong Excel File
            self.message_gen_final.changeText('Generating a RCA Schedule Fails!! - Missing Excel File')
            self.message_gen_final.changeColor('red')

    ################################################################ Class Function
    # Class Function to click to text a newly generated RCA schedule
    def Click_Txt_Sch(self):
        rS = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\/\d{1,2}\/\d{1,2}-\d{1,2}\/\d{1,2}\/\d{1,2}\)')
        rSA = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\.\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}\.\d{1,2}\)')
        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
        if self.resultFlag[1] != "":                    # Excel File Check
            # self.resultFlag[0] = "gen"
            if (((rS.match(str(self.entry_gen_forsheet.getText()).replace(" ","").replace("\n", "")) is not None 
                    or rSA.match(str(self.entry_gen_forsheet.getText()).replace(" ","").replace("\n", "")) is not None)
                        and self.Days_Between(str(self.entry_gen_forsheet.getText()).replace(" ","").replace("\n", ""))==6)
                            or self.resultFlag[2] != ""):              # Sheet Name Check
                if (rS.match(str(self.entry_gen_forsheet.getText()).strip().replace(" ","").lower()) is not None
                    or rSA.match(str(self.entry_gen_forsheet.getText()).strip().replace(" ","").lower()) is not None):
                    # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                    self.resultFlag[2]= str(self.entry_gen_forsheet.getText()).strip().replace(" ","").lower()
                # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"                  
                self.resultFlag[3] = TRUE

                # ################## Start Working on Excel File
                returned_Value_FromExcel = self.Working_With_Excel()

                if returned_Value_FromExcel[0]:
                    temp_Sheet_Name = self.resultFlag[2].replace('/', '.')
                    temp_Msg = returned_Value_FromExcel[1] + ' - ' + temp_Sheet_Name
                    self.message_gen_final.changeText(temp_Msg)
                    self.message_gen_final.changeColor('green')
                else:
                    self.message_gen_final.changeText(returned_Value_FromExcel[1])
                    self.message_gen_final.changeColor('red')
            else:
                # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                self.resultFlag[3] = FALSE
                self.message_gen_final.changeText('Texting a RCA Schedule Fails!! - Missing or Wrong Sheet\'s Name')
                self.message_gen_final.changeColor('red')
        else:
            # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
            self.resultFlag[3] = FALSE
            self.message_gen_final.changeText('Texting a RCA Schedule Fails!! - Excel File Missing')
            self.message_gen_final.changeColor('red')     
    
    ####################################################################################### Class Function
    # Class Function to click to read in the contents of a sheet (an existing RCA schedule)
    def Click_Get_SheetName(self):
        rS = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\/\d{1,2}\/\d{1,2}-\d{1,2}\/\d{1,2}\/\d{1,2}\)')
        rSA = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\.\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}\.\d{1,2}\)')
        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
        if self.resultFlag[1] !=  '':
            if ((rS.match(str(self.entry_mod_forsheet.getText()).strip().replace(" ","")) is not None
                or rSA.match(str(self.entry_mod_forsheet.getText()).strip().replace(" ","")) is not None) 
                    and self.Days_Between(str(self.entry_mod_forsheet.getText()).replace(" ","").replace("\n", ""))==6):
                # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                # self.resultFlag[0] = "mod"
                self.resultFlag[2] = str(self.entry_mod_forsheet.getText()).strip().replace(" ","").lower()
                self.resultFlag[3] = FALSE
                self.resultFlag[4] = []
                self.message_mod_final.changeText('')

                # ################## Start Working on Excel File
                returned_Value_FromExcel = self.Working_With_Excel()

                if returned_Value_FromExcel[0]:
                    self.message_mod_final.changeText(returned_Value_FromExcel[1])
                    self.message_mod_final.changeColor('green')
                else:
                    self.message_mod_final.changeText(returned_Value_FromExcel[1])
                    self.message_mod_final.changeColor('red')
            else:
                self.message_mod_final.changeText('Reading in Contents Fails!! - Missing or Wrong Sheet\'s Name')
                self.message_mod_final.changeColor('red')
        else:
            self.message_mod_final.changeText('Reading in Contents Fails!! - Missing Excel File')
            self.message_mod_final.changeColor('red')
            
    ############################################################## Class Function
    # Class Function to click to text an updated RCA schedule only
    def Click_Txt_Update(self):
        rS = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\/\d{1,2}\/\d{1,2}-\d{1,2}\/\d{1,2}\/\d{1,2}\)')
        rSA = re.compile('ww\d{1,2}\.[1,2]{1}\(\d{1,2}\.\d{1,2}\.\d{1,2}-\d{1,2}\.\d{1,2}\.\d{1,2}\)')
        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
        if self.resultFlag[0] == 'mod':                             # Update Mode Check
            if self.resultFlag[1] !=  '':                           # Excel File Check
                if (rS.match(self.resultFlag[2]) is not None or rSA.match(self.resultFlag[2]) is not None):        # Sheet Name Check      # Success
                    # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                    self.resultFlag[3] = TRUE
                    self.resultFlag[4] = []

                    # ################## Start Working on Excel File
                    returned_Value_FromExcel = self.Working_With_Excel()

                    if returned_Value_FromExcel[0]:
                        temp_Sheet_Name = self.resultFlag[2].replace('/', '.')
                        temp_Msg = returned_Value_FromExcel[1] + ' - Updates in ' + temp_Sheet_Name
                        self.message_mod_final.changeText(temp_Msg)
                        self.message_mod_final.changeColor('green')
                    else:
                        self.message_mod_final.changeText(returned_Value_FromExcel[1])
                        self.message_mod_final.changeColor('red')
                else:                                               # Fail
                    # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                    self.resultFlag[3] = FALSE
                    self.resultFlag[4] = []
                    self.message_mod_final.changeText('Texting an Update Fails!! - Missing or Wrong Sheet\'s Name')
                    self.message_mod_final.changeColor('red')
            else:                                                   # Fail
                # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
                self.resultFlag[3] = FALSE
                self.resultFlag[4] = []
                self.message_mod_final.changeText('Texting an Update Fails!! - Wrong Excel File')
                self.message_mod_final.changeColor('red')
        else:                                                       # Fail
            # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
            self.resultFlag[3] = FALSE
            self.resultFlag[4] = []
            self.message_mod_final.changeText('Texting an Update Fails!! - No Update Mode')
            self.message_mod_final.changeColor('red')

    ####################################################################################################### Class Function
    # Class Function to click to open up a sub-window to set up some parameters for RCA schedule generation
    def Click_Setup(self, masterS):
            newWindowS = tk.Toplevel(masterS)
            newWindowS.resizable(False, False)

            newWindowS.grid_columnconfigure(0, weight=1, uniform="fred")
            newWindowS.grid_columnconfigure(1, weight=1, uniform="fred")
            newWindowS.grid_columnconfigure(2, weight=1, uniform="fred")
            newWindowS.title("Set Up")
        
            # sets the geometry of toplevel
            newWindowS.geometry("500x460")

            # 1st Label in Set Up
            self.label_setup00 = label_gui(newWindowS)
            self.label_setup00.setLabel('RCA Scheduler SetUp', 'Helvetica 10 bold', 10)
            self.label_setup00.setGridLabel(0, 0, 3)

            # Setting for Printing Label in set Up
            self.label_setup10 = label_gui(newWindowS)
            self.label_setup10.setLabel('--- Setting for Printing ---', 'Helvetica 9 bold', 5)
            self.label_setup10.setGridLabel(0, 1, 3)
        
            # Label for Setting for Printing - Cell Column Width
            self.label_setup20 = label_gui(newWindowS)
            # self.label_setup20.setLabel0('File Path: ', 'Helvetica 10 bold', 20, 5, LEFT, 5, 'blue')
            self.label_setup20.setLabel('Cell Column Width:', 'Helvetica 9', 6)
            self.label_setup20.setGridLabel(0, 2, 1)
            self.label_setup20.configurePadx(15)

            # Entry for Setting for Printing - Cell Column Width
            self.entry_setup21 = entry_gui(newWindowS)
            self.entry_setup21.setEntry('', 24, 3, 'light gray')
            self.entry_setup21.setGridEntry(1, 2, 2, 0)
            self.entry_setup21.deleteValueEntry()
            self.entry_setup21.setValueEntry("0")
            self.entry_setup21.setSticky('W')
            self.entry_setup21.setJustify('right')
            self.entry_setup21.setPadx(6)

            # Label for Default Value of Cell Column Width
            self.label_setup22 = label_gui(newWindowS)
            self.label_setup22.setLabel('14 - by default, (> 6)', 'Helvetica 9', 6)
            self.label_setup22.setFontColor('orange')
            self.label_setup22.setGridLabel(2, 2, 1)
            self.label_setup22.configurePadx(0)
            self.label_setup22.setSticky('W')
            self.label_setup22.setPadx(4)

            # Button for Text Schedule in Generator
            self.button_setup50 = button_gui(newWindowS)
            self.button_setup50.setButton('Apply', 10, "gray", 5, "yellow", lambda:self.Click_Cell_Width_Length_Setting_Apply(newWindowS))
            self.button_setup50.setGridButton(2, 4, 1, 0)
            # self.button_setup91.setSticky('W')
            self.button_setup50.setPadx(6)
            self.button_setup50.setPady(6)

            ################################################ 2
            # SMS Texting Label in set Up
            self.label_setup60 = label_gui(newWindowS)
            self.label_setup60.setLabel('--- Settings for RCA Schedule Generation ---', 'Helvetica 9 bold', 8)
            self.label_setup60.setGridLabel(0, 5, 3)

            #####
            # Label for New RTE Insertion
            self.label_setup70 = label_gui(newWindowS)
            self.label_setup70.setLabel('New RTE Insertion Rate:', 'Helvetica 9', 6)
            self.label_setup70.setGridLabel(0, 6, 1)
            self.label_setup70.configurePadx(15)

            # Entry for New RTE Insertion
            self.entry_setup71 = entry_gui(newWindowS)
            self.entry_setup71.setEntry('', 24, 3, 'light gray')
            self.entry_setup71.setGridEntry(1, 6, 1, 0)
            self.entry_setup71.deleteValueEntry()
            self.entry_setup71.setValueEntry("0")
            self.entry_setup71.setSticky('W')
            self.entry_setup71.setJustify('right')
            self.entry_setup71.setPadx(6)

            # Label for New RTE Insertion (%)
            self.label_setup72 = label_gui(newWindowS)
            self.label_setup72.setLabel('% (0 - 100)', 'Helvetica 9', 6)
            self.label_setup72.setFontColor('orange')
            self.label_setup72.setGridLabel(2, 6, 1)
            self.label_setup72.configurePadx(0)
            self.label_setup72.setSticky('W')
            self.label_setup72.setPadx(4)
            # self.label_setup31.setJustify('left')
            ##### End

            #####
            # Label for Off-Day Requirement
            self.label_setup80 = label_gui(newWindowS)
            self.label_setup80.setLabel('Is One Off-Day Required?:', 'Helvetica 9', 6)
            self.label_setup80.setGridLabel(0, 7, 1)
            self.label_setup80.configurePadx(15)

            # Radio Button for Off-Day Requirement
            self.var_One_Off_Day_Req = tk.StringVar(value='No')
            self.radio_button_setup81 = radiobutton_gui(newWindowS)
            self.radio_button_setup81.setRadioButton('Yes', self.var_One_Off_Day_Req, 'Yes')
            self.radio_button_setup81.setGridRadioButton(1, 7, 1)
            self.radio_button_setup81.setPadyRadioButton(6)
            # self.radio_button_setup81.setStickyRadioButton('W')
            self.radio_button_setup82 = radiobutton_gui(newWindowS)
            self.radio_button_setup82.setRadioButton('No (by default)', self.var_One_Off_Day_Req, 'No')
            self.radio_button_setup82.setGridRadioButton(2, 7, 1)
            self.radio_button_setup82.setStickyRadioButton('W')
            self.radio_button_setup82.setPadyRadioButton(6)

            ##### End

            # Button for Text Schedule in Generator
            self.button_setup90 = button_gui(newWindowS)
            self.button_setup90.setButton('Apply', 10, "gray", 5, "yellow", lambda:self.Click_RCA_Sch_Gen_Setting_Apply())
            self.button_setup90.setGridButton(2, 8, 1, 0)
            # self.button_setup91.setSticky('E')
            self.button_setup90.setPadx(6)
            self.button_setup90.setPady(6)

            # Title - Supervisor Info set Up
            self.label_setup100 = label_gui(newWindowS)
            self.label_setup100.setLabel('--- Setting to Check Texting\'s Success ---', 'Helvetica 9 bold', 8)
            self.label_setup100.setGridLabel(0, 9, 3)

            # Label for Supervisor Info
            self.label_setup110 = label_gui(newWindowS)
            self.label_setup110.setLabel('Supervisor Phone #:', 'Helvetica 9', 6)
            self.label_setup110.setGridLabel(0, 10, 1)
            self.label_setup110.configurePadx(15)

            # Entry for New RTE Insertion
            self.entry_setup111 = entry_gui(newWindowS)
            self.entry_setup111.setEntry('', 24, 3, 'light gray')
            self.entry_setup111.setGridEntry(1, 10, 1, 0)
            self.entry_setup111.deleteValueEntry()
            self.entry_setup111.setValueEntry("000-000-0000")
            self.entry_setup111.setSticky('W')
            self.entry_setup111.setPadx(6)
        
            # Button for Text Schedule in Generator
            self.button_setup120 = button_gui(newWindowS)
            self.button_setup120.setButton('Apply', 10, "gray", 5, "yellow", lambda:self.Click_Sup_Phone_Setting_Apply())
            self.button_setup120.setGridButton(2, 11, 1, 0)
            # self.button_setup91.setSticky('E')
            self.button_setup120.setPadx(6)
            self.button_setup120.setPady(6)

            # Button for Text Schedule in Generator
            self.button_setup130 = button_gui(newWindowS)
            self.button_setup130.setButton('CLOSE', 10, "white", 5, "black", lambda:self.Click_Close(newWindowS))
            self.button_setup130.setGridButton(2, 12, 1, 0)
            # self.button_setup91.setSticky('W')
            self.button_setup130.setPadx(6)
            self.button_setup130.setPady(6)

            newWindowS.transient(masterS)
            newWindowS.grab_set()
            masterS.wait_window(newWindowS)

    ####################################################################################### Class Function
    # Class Function to click to apply setting for cell column with to fit for priting size
    def Click_Cell_Width_Length_Setting_Apply(self, wb):

        if str(self.entry_setup21.getText()).strip().replace(' ', '') != '0':
            temp_CellColWidth = str(self.entry_setup21.getText()).strip().replace(' ', '')
            if temp_CellColWidth.isdigit() and len(temp_CellColWidth) <= 2:
                NewCellColWidth = temp_CellColWidth
            else:
                NewCellColWidth = 16
        else:
            NewCellColWidth = 16
        self.myExcel.Set_New_Cell_Column_Width(NewCellColWidth)

    ########################################################################################### Class Function
    # Class Function to click to apply settings that users provides for RCA Schedule Generation 
    def Click_RCA_Sch_Gen_Setting_Apply(self):
        
        # Verify Inputs of New Route Insertion Rate
        if str(self.entry_setup71.getText()).strip().replace(' ', '') != '0':
            temp_SMS_NewRTERate = str(self.entry_setup71.getText()).strip().replace(' ', '')
            temp_Integer1 = int(temp_SMS_NewRTERate)
            if temp_SMS_NewRTERate.isdigit() and len(temp_SMS_NewRTERate) <= 3 and (temp_Integer1 <= 100 and temp_Integer1 > 0):
                NewRTERate = temp_SMS_NewRTERate
            else:
                NewRTERate = 0
        else:
            NewRTERate = 0

        self.myExcel.Set_New_Route_Insertion_Rate(NewRTERate)
        self.myExcel.Set_One_OffDay_Required(self.var_One_Off_Day_Req.get())
        

    ################################################################### Class Function
    # Class Function to click to apply Supervisor Phone Number to check 
    #                if texting of RCA's schedule is successfully texted to RCA
    def Click_Sup_Phone_Setting_Apply(self):

        # Verify Inputs of Supervisor Phone
        if str(self.entry_setup111.getText()).strip().replace(' ', '') != '000-000-0000':
            temp_SUPR_PhoneS = str(self.entry_setup111.getText()).strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('.', '')
            if temp_SUPR_PhoneS.isdigit() and len(temp_SUPR_PhoneS) == 10:
                SUPR_PhoneS = '+1' + temp_SUPR_PhoneS
            else:
                SUPR_PhoneS = None
        else:
            SUPR_PhoneS = None

        ### Apply settings into environment variables
        self.myExcel.Set_Supervisor_Phone(SUPR_PhoneS)
        

    ###################################################################################### Class Function
    # Class Function to work with Excel to create scheduler and modify an existin schedule - Main Big Part
    def Working_With_Excel(self):
    
        return_Msg = [FALSE, '']
        flag_OffDays = [str(item).lower() for item in self.resultFlag[4]]
        if self.resultFlag[1] is not None:                            # Check if an excel file exists
            wb=load_workbook(self.resultFlag[1])
            self.myExcel.set_Wb(wb)
        else:
            return_Msg[0] = FALSE
            return_Msg[1] = 'Missing Excel File!!'

        # Execution based on Button Actions
        # "gen or mod", excel file name, sheet name, TRUE or FALSE for text, "[off-day]"
        if self.resultFlag[0] == 'gen':                                           # 'Gen' Mode
            if self.resultFlag[3] == TRUE:                        # Texting Generated Schedule
                if self.resultFlag[2] is not None:                # Check if sheet name exists
                    self.myExcel.Empty_A_Variable_RCA_Name()
                    self.myExcel.Empty_A_Variable_RCA_Phone()
                    self.myExcel.Empty_A_Variable_RCA_Value()
                    self.myExcel.Empty_A_Variable_RCA_PreviousRoute_Info()
                    self.myExcel.Empty_A_Variable_Route_Name()
                    self.myExcel.Empty_A_Variable_Route_Value()
                    self.myExcel.Read_RCA()
                    self.myExcel.Read_Route()
                    if not self.myExcel.Verify_A_Sheet_Existed_Or_After_Created(self.resultFlag[2], self.resultFlag[1]):
                        return_Msg[0] = FALSE
                        return_Msg[1] = 'This Schedule Not Existed!!'
                        return return_Msg
                    self.myExcel.Texting_New_Sch_Or_Updated(self.resultFlag[0], self.resultFlag[2], self.myExcel.Read_A_Sheet_Of_Existing_RCA_Schedule_Or_Just_Created_RCA_Schedule(self.resultFlag[2], self.resultFlag[1])[1])
                    return_Msg[0] = TRUE
                    return_Msg[1] = 'A RCA Schedule is Successfully Texted!!'
                    return return_Msg
                    
                else:                                   # Failed - missing sheet name
                    return_Msg[0] = FALSE
                    return_Msg[1] = 'Missing Sheet Name!!'
                    return return_Msg
            else:                                       # Pure Gen
                self.myExcel.Empty_All_Variables()
                if self.myExcel.Verify_A_Sheet_Existed_Or_After_Created(self.resultFlag[2], None):       # Check before RCA schedule generation
                    return_Msg[0] = FALSE
                    return_Msg[1] = 'This Schedule Already Existed!!'
                    return return_Msg
                self.myExcel.Read_RCA()
                self.myExcel.Read_Route()
                self.myExcel.Read_Multi_Sheets()
                self.myExcel.Create_Scheduler(flag_OffDays)                                # with Off-days list
                self.myExcel.Write_Generated_Sch_2Sheet(self.resultFlag[2], self.resultFlag[1])                # with sheet name
                self.myExcel.Adjust_Size_Of_Cells(self.resultFlag[2], self.resultFlag[1])
                if self.myExcel.Verify_A_Sheet_Existed_Or_After_Created(self.resultFlag[2], self.resultFlag[1]):
                    return_Msg[0] = TRUE
                    return_Msg[1] = 'A RCA Schedule is Successfully Generated!!'
                    return return_Msg
                else:
                    return_Msg[0] = FALSE
                    return_Msg[1] = 'Generating a RCA Schedule Fails!!'
                    return return_Msg

        ###### Modified
        else:                                                           # 'Mod' Mode
            if self.resultFlag[2] is not None:                                    # Read in a sheet contents
                if self.resultFlag[3] ==  FALSE:
                    self.myExcel.Empty_A_Variable_Sch_1stRead()

                    if not self.myExcel.Verify_A_Sheet_Existed_Or_After_Created(self.resultFlag[2], None):
                        return_Msg[0] = FALSE
                        return_Msg[1] = 'This Schedule Not Existed!!'
                        return return_Msg
                    self.myExcel.Read_A_Sheet_1st(self.resultFlag[2])
                    # self.myExcel.Make_A_Same_Sheet_4Update(self.resultFlag[2], self.resultFlag[1])
                    # self.myExcel.Adjust_Size_Of_Cells(self.resultFlag[2], self.resultFlag[1])
                    return_Msg[0] = TRUE
                    return_Msg[1] = 'Contents is Successfully Read In!!'
                    return return_Msg
                else:                                                   # Text only updated
                    self.myExcel.Empty_A_Variable_RCA_Name()
                    self.myExcel.Empty_A_Variable_RCA_Phone()
                    self.myExcel.Empty_A_Variable_RCA_Value()

                    self.myExcel.Read_RCA()

                    self.myExcel.Empty_A_Variable_Sch_2ndRead_AfterModifying()
                    if not self.myExcel.Verify_A_Sheet_Existed_Or_After_Created(self.resultFlag[2], None):
                        return_Msg[0] = FALSE
                        return_Msg[1] = 'This Schedule Not Existed!!'
                        return return_Msg
                    self.myExcel.Read_A_Sheet_2nd(self.resultFlag[2])                      # read a sheet content again
                    self.myExcel.Texting_New_Sch_Or_Updated(self.resultFlag[0], self.resultFlag[2], [])
                    return_Msg[0] = TRUE
                    return_Msg[1] = 'Update is Successfully Texted!!'
                    return return_Msg
            else:
                return_Msg[0] = FALSE
                return_Msg[1] = 'Missing Sheet Name!!'
                return return_Msg
       

    #################################################### Class Function
    # Class Functio to click to close sub 'setup' window
    def Click_Close(self, newWindowSS):
        self.button_setup.setState('normal')
        newWindowSS.destroy()

    #################################################### Class Function
    # Class Functio to click to close main window
    def Click_Close_Main(self, masterMM):
        self.button_setup.setState('normal')
        masterMM.destroy()

    #################################################### Class Function - Support

    # 1. Check the format of off-days
    def Check_Format_ListsOfOffDays(self, listS):
        listDays = ['sat', 'mon', 'tue', 'wed', 'thu', 'fri']
        for i in range (0, len(listS)):
            x = FALSE
            for j in range (0, len(listDays)):
                if str(listS[i]) == listDays[j]:
                    x=TRUE
                    break

            if x == FALSE:
                return FALSE
        return TRUE

    # 2. Check the difference between two dates
    def Days_Between(self, sheetNames):
        twoDates = self.Convert_SheetName_Dates(sheetNames)
        d10 = self.Change_Format_Date(twoDates[0])
        d20 = self.Change_Format_Date(twoDates[1])
        # Year Change
        if d10[0] != d20[0]:                    # Year Change
            return (31 - d10[2]) + d20[2]
        else:                                   # Year Not Change
            if d10[1] != d20[1]:                    # Month Change
                if d10[1] == 2 and self.myExcel.isAMultipleOf4(d10[0]):        # Feb == 29
                    return (29 - d10[2]) + d20[2]
                elif d10[1] == 2 and self.myExcel.isAMultipleOf4(d10[0]):        # Feb == 28
                    return (28 - d10[2]) + d20[2]
                elif d10[1] == 4 or d10[1] == 6 or d10[1] == 9 or d10[1] == 11:   # Month with 30 days
                    return (30 - d10[2]) + d20[2]
                else:                                                               # Month with 31 days
                    return (31 - d10[2]) + d20[2]
            else:                                   # Month Not Change
                return d20[2] - d10[2]

    # 3. Change the format of date
    def Change_Format_Date(self, dateS):
        if '/' in dateS:
            tempDate = dateS.split('/')
        if '.' in dateS:
            tempDate = dateS.split('.')
        return [int('20' + tempDate[2]), int(tempDate[0]), int(tempDate[1])]

    # 4. Convert Sheet's Name into two Dates
    def Convert_SheetName_Dates(self, sheetNameS):
        tempStr = sheetNameS.split('(')
        tempStr10 = tempStr[1].split('-')
        firstDate = tempStr10[0]                                # first date
        tempStr21 = tempStr10[1].split(')')
        secondDate = tempStr21[0]                               # second date
        return [firstDate, secondDate]


#
############################ GUI Applicaiton
#
### Global Function 'App_GUI'
def App_GUI():
    root = tk.Tk()
    root.resizable(False, False)
    root.title('RCA Weekly Scheduler')
    root.geometry('900x580')
    myRCA_Scheduler_And_Modifier = RCA_Scheduler_Modifier(root)
   
    root.mainloop()


# 
# ########### Main ############
# 
if __name__ == '__main__':
    # if not process_exists('rca_sch.exe'):
    App_GUI()
        